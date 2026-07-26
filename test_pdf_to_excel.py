#!/usr/bin/env python3
"""
Test suite for pdf_to_excel.py
Generates synthetic test PDFs and validates output Excel files.
"""

import os
import re
import sys
import json
import argparse
import logging
import unittest
import unittest.mock
import tempfile
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# Allow importing from parent directory
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# ── Optional PDF generation for test fixtures ─────────────────────────────────
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    )
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

from pdf_to_excel import (
    load_config, setup_logging, PDFClassifier, TableExtractor,
    TextExtractor, ReportColumnParser, ExcelBuilder, PDFProcessor,
    BatchProcessor, check_dependencies, export_to_csv
)


def make_test_dir() -> str:
    """A scratch directory under the system temp dir.

    Tests used to write into a fixed folder in the user's home directory, which
    left a directory behind on every machine that ever ran the suite. The system
    temp dir is what the platform provides for exactly this, and each test's
    tearDown removes its own tree.
    """
    return tempfile.mkdtemp(prefix="pdf2xl_test_")


def close_logger(logger: logging.Logger) -> None:
    for handler in list(logger.handlers):
        handler.close()
        logger.removeHandler(handler)


# ─────────────────────────────────────────────────────────────────────────────
# FIXTURES
# ─────────────────────────────────────────────────────────────────────────────

def create_test_pdf_with_table(path: str) -> None:
    if not REPORTLAB_AVAILABLE:
        raise unittest.SkipTest("reportlab not installed; skipping PDF generation tests")

    doc = SimpleDocTemplate(path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Sales Report Q1 2024", styles["Title"]))
    elements.append(Spacer(1, 12))

    data = [
        ["Product", "Units Sold", "Revenue ($)", "Region"],
        ["Widget A", "1,200", "24,000.00", "North"],
        ["Widget B", "850", "17,000.00", "South"],
        ["Gadget X", "2,100", "63,000.00", "East"],
        ["Gadget Y", "430", "12,900.00", "West"],
        ["Super Pro", "99", "49,500.00", "North"],
    ]

    t = Table(data, colWidths=[120, 80, 100, 80])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))
    elements.append(t)
    doc.build(elements)


def create_multi_table_pdf(path: str) -> None:
    if not REPORTLAB_AVAILABLE:
        raise unittest.SkipTest("reportlab not installed")

    doc = SimpleDocTemplate(path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    for section_num in range(1, 4):
        elements.append(Paragraph(f"Section {section_num}", styles["Heading1"]))
        elements.append(Spacer(1, 8))

        data = [["ID", "Name", "Value"]] + [
            [str(i), f"Item {i}", str(i * 100)]
            for i in range(1, 6)
        ]
        t = Table(data, colWidths=[60, 150, 80])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.navy),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 20))

    doc.build(elements)


def create_text_only_pdf(path: str) -> None:
    if not REPORTLAB_AVAILABLE:
        raise unittest.SkipTest("reportlab not installed")

    doc = SimpleDocTemplate(path, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Annual Report 2024", styles["Title"]),
        Spacer(1, 12),
        Paragraph(
            "This document contains our annual performance summary. "
            "Revenue grew 15% year over year driven by strong product demand.",
            styles["Normal"]
        ),
        Spacer(1, 8),
        Paragraph(
            "Key highlights: 500 new customers, 98% retention rate, "
            "12 new product launches across 4 markets.",
            styles["Normal"]
        ),
    ]
    doc.build(elements)


def create_order_report_pdf(path: str) -> None:
    """A supplier order report: no ruled grid, values in real columns.

    Drawn with explicit x positions rather than as a platypus Table, because
    the point of the fixture is the column *geometry* - the report carries no
    ruling lines, so the table engines find nothing and the column parser is
    what has to read it. That is exactly the document CSV export used to miss.
    """
    if not REPORTLAB_AVAILABLE:
        raise unittest.SkipTest("reportlab not installed")

    from reportlab.pdfgen import canvas as pdfcanvas

    columns = [
        ("S. No.", 40), ("Customer Name", 80), ("Item", 170),
        ("Product Type", 260), ("Certificate No", 330), ("Order No", 410),
        ("Amount", 470), ("Manufacturer", 530), ("Est Shipping Date", 620),
    ]
    rows = [
        ["1", "Alex Morgan", "2.14 Round J VS1", "Diamond", "IGI",
         "A1B2C3D4", "$413.25", "Guild and Facet", "27 Apr 2026"],
        ["2", "Jordan Avery", "2.20 Emerald J", "Diamond", "IGI 900000001",
         "E5F6A7B8", "$4722.30", "Guild and Facet", "24 Apr 2026"],
    ]

    c = pdfcanvas.Canvas(path, pagesize=(792, 612))
    c.setFont("Helvetica", 9)
    c.drawString(40, 560, "Invoice #GF2026042110")
    c.drawString(40, 546, "2 Items - $5135.55")

    y = 520
    for label, x in columns:
        c.drawString(x, y, label)
    for row in rows:
        y -= 18
        for (_label, x), value in zip(columns, row):
            c.drawString(x, y, value)
    c.showPage()
    c.save()


def create_table_then_report_pdf(path: str) -> None:
    """Page 1: a ruled table. Page 2: an order report with no ruling.

    The mixed case that used to lose data - one table anywhere in the document
    stopped the text/geometry path from running at all, so page 2 was never
    read and its excluded-row accounting never happened.
    """
    if not REPORTLAB_AVAILABLE:
        raise unittest.SkipTest("reportlab not installed")

    from reportlab.pdfgen import canvas as pdfcanvas

    c = pdfcanvas.Canvas(path, pagesize=(792, 612))

    # -- Page 1: a genuinely ruled grid, so pdfplumber finds a table ----------
    # Enough prose to clear the classifier's >50-character text threshold;
    # below it the page is treated as scanned and never reaches a table engine,
    # which would make this fixture prove nothing.
    c.setFont("Helvetica", 10)
    c.drawString(60, 560, "Quarterly regional summary for the northern and southern")
    c.drawString(60, 546, "territories, reported in units shipped during the period.")

    grid = [
        ["Region", "Units", "Revenue"],
        ["North", "120", "24000.00"],
        ["South", "340", "68000.00"],
        ["East", "215", "43000.00"],
    ]
    x0, y0, cw, rh = 60, 500, 120, 24
    for r, row in enumerate(grid):
        for col, value in enumerate(row):
            c.rect(x0 + col * cw, y0 - r * rh, cw, rh)
            c.drawString(x0 + col * cw + 6, y0 - r * rh + 8, value)
    c.showPage()

    # -- Page 2: the order report, drawn in columns with no ruling -----------
    c.setFont("Helvetica", 9)
    columns = [
        ("S. No.", 40), ("Customer Name", 80), ("Item", 170),
        ("Product Type", 260), ("Certificate No", 330), ("Order No", 410),
        ("Amount", 470), ("Manufacturer", 530), ("Est Shipping Date", 620),
    ]
    rows = [
        ["1", "Alex Morgan", "2.14 Round J VS1", "Diamond", "IGI",
         "A1B2C3D4", "$413.25", "Guild and Facet", "27 Apr 2026"],
        ["2", "Jordan Avery", "2.20 Emerald J", "Diamond", "IGI 900000001",
         "E5F6A7B8", "$4722.30", "Guild and Facet", "24 Apr 2026"],
    ]
    y = 520
    for label, x in columns:
        c.drawString(x, y, label)
    for row in rows:
        y -= 18
        for (_label, x), value in zip(columns, row):
            c.drawString(x, y, value)
    c.showPage()
    c.save()


REPORT_COLUMNS = [
    "S. No.", "Customer Name", "Item", "Product Type", "Certificate No",
    "Order No", "Amount", "Manufacturer", "Est Shipping Date",
]


# ─────────────────────────────────────────────────────────────────────────────
# TEST CASES
# ─────────────────────────────────────────────────────────────────────────────

class TestDependencies(unittest.TestCase):
    def test_check_dependencies_returns_dict(self):
        deps = check_dependencies()
        self.assertIsInstance(deps, dict)
        self.assertIn("pdfplumber", deps)
        self.assertIn("PyMuPDF (fitz)", deps)
        self.assertIn("pandas", deps)
        self.assertIn("openpyxl", deps)

    def test_core_deps_available(self):
        deps = check_dependencies()
        self.assertTrue(deps["pandas"], "pandas must be installed")
        self.assertTrue(deps["openpyxl"], "openpyxl must be installed")


class TestDocsMatchCode(unittest.TestCase):
    """The README is the single reference surface, so it must not go stale.

    index.html used to restate the CLI flags, the config keys, installation and
    troubleshooting in full - roughly 720 duplicated lines that had to be edited
    in step with the README and, predictably, did not stay in step. It now links
    here instead. That removes one copy; these tests stop the remaining one from
    drifting away from the code it describes.
    """

    ROOT = os.path.dirname(os.path.abspath(__file__))

    @classmethod
    def setUpClass(cls):
        with open(os.path.join(cls.ROOT, "README.md"), encoding="utf-8") as fh:
            cls.readme = fh.read()

    def _read(self, name: str) -> str:
        with open(os.path.join(self.ROOT, name), encoding="utf-8") as fh:
            return fh.read()

    # ── CLI ──────────────────────────────────────────────────────────────
    def test_every_cli_flag_is_documented(self):
        from pdf_to_excel import build_parser

        flags = {
            opt
            for action in build_parser()._actions
            # argparse supplies --help itself and the usage line already shows
            # -h; documenting it separately would be noise.
            if not isinstance(action, argparse._HelpAction)
            for opt in action.option_strings
            if opt.startswith("--")
        }
        missing = sorted(f for f in flags if f not in self.readme)
        self.assertEqual(
            missing, [],
            f"CLI flags exist in build_parser() but are absent from README.md: {missing}"
        )

    def test_readme_documents_no_flag_that_does_not_exist(self):
        from pdf_to_excel import build_parser

        real = {
            opt
            for action in build_parser()._actions
            for opt in action.option_strings
        }
        # Scoped to the fenced usage block so ordinary prose cannot trip it.
        block = re.search(
            r"## CLI Reference\s*```(.*?)```", self.readme, re.S
        )
        self.assertIsNotNone(block, "README.md has no fenced CLI Reference block")
        documented = set(re.findall(r"--[a-z][a-z0-9-]*", block.group(1)))
        phantom = sorted(documented - real)
        self.assertEqual(
            phantom, [],
            f"README.md documents flags the parser does not define: {phantom}"
        )

    # ── Configuration ────────────────────────────────────────────────────
    def test_every_config_key_is_documented(self):
        from pdf_to_excel import DEFAULT_CONFIG

        missing = sorted(k for k in DEFAULT_CONFIG if k not in self.readme)
        self.assertEqual(
            missing, [],
            f"config keys in DEFAULT_CONFIG but undocumented in README.md: {missing}"
        )

    def test_shipped_config_defines_no_unknown_keys(self):
        from pdf_to_excel import DEFAULT_CONFIG

        shipped = json.loads(self._read("config.json"))
        # Keys prefixed with _ are commentary, not settings.
        unknown = sorted(
            k for k in shipped
            if not k.startswith("_") and k not in DEFAULT_CONFIG
        )
        self.assertEqual(
            unknown, [],
            f"config.json sets keys the code never reads: {unknown}"
        )

    # ── The consolidation itself ─────────────────────────────────────────
    def test_index_html_points_at_the_readme_instead_of_copying_it(self):
        html = self._read("index.html")
        for duplicated in ('id="install"', 'id="usage"', 'id="cli"',
                           'id="config"', 'id="output"', 'id="troubleshoot"'):
            self.assertNotIn(
                duplicated, html,
                f"index.html has re-grown a section that duplicates the README "
                f"({duplicated}). Reference material belongs in README.md."
            )
        self.assertIn('id="reference"', html)
        self.assertIn("README.md", html)

    def test_readme_screenshot_exists(self):
        # A README whose lead image 404s looks worse than one with no image.
        for match in re.findall(r"!\[[^\]]*\]\(([^)]+)\)", self.readme):
            if match.startswith("http"):
                continue
            path = os.path.join(self.ROOT, match.replace("/", os.sep))
            self.assertTrue(
                os.path.isfile(path),
                f"README.md references an image that is not in the repo: {match}"
            )

    def test_stated_test_count_is_accurate(self):
        # Both README.md and PRODUCT.md advertise this number, and it has gone
        # stale twice. It is a claim made to anyone reading the repo.
        actual = unittest.TestLoader().loadTestsFromModule(
            sys.modules[__name__]
        ).countTestCases()

        for name in ("README.md", "PRODUCT.md"):
            stated = set(re.findall(r"(\d+)\s+(?:tests|%20passing)", self._read(name)))
            stated |= set(re.findall(r"tests-(\d+)%20passing", self._read(name)))
            self.assertTrue(
                stated, f"{name} states no test count"
            )
            self.assertEqual(
                stated, {str(actual)},
                f"{name} states test count {sorted(stated)} but the suite has "
                f"{actual}. Update the number in README.md and PRODUCT.md."
            )


class TestConfig(unittest.TestCase):
    def setUp(self):
        self.tmp = make_test_dir()

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def test_load_defaults(self):
        cfg = load_config("nonexistent_config.json")
        self.assertEqual(cfg["ocr_language"], "eng")
        self.assertTrue(cfg["remove_empty_rows"])
        self.assertTrue(cfg["add_metadata_sheet"])

    def test_load_custom_config(self):
        cfg_path = os.path.join(self.tmp, "test_config.json")
        custom = {"ocr_language": "fra", "ocr_dpi": 150, "header_color": "FF0000"}
        with open(cfg_path, "w") as f:
            json.dump(custom, f)
        cfg = load_config(cfg_path)
        self.assertEqual(cfg["ocr_language"], "fra")
        self.assertEqual(cfg["ocr_dpi"], 150)
        self.assertEqual(cfg["header_color"], "FF0000")
        # Defaults preserved for unset keys
        self.assertTrue(cfg["remove_empty_rows"])

    def test_malformed_config_uses_defaults(self):
        cfg_path = os.path.join(self.tmp, "bad_config.json")
        with open(cfg_path, "w") as f:
            f.write("{invalid json}")
        cfg = load_config(cfg_path)
        self.assertEqual(cfg["ocr_language"], "eng")


class TestOptionalDependencyHandling(unittest.TestCase):
    def test_tesseract_path_without_pytesseract_does_not_raise(self):
        # pytesseract is an optional import, so reading its attributes when the
        # package is absent raised NameError at construction - hitting exactly
        # the users who had followed the README and set tesseract_path.
        import pdf_to_excel as module
        from pdf_to_excel import OCRProcessor

        original = module.TESSERACT_AVAILABLE
        module.TESSERACT_AVAILABLE = False
        try:
            OCRProcessor(
                {"tesseract_path": "C:/nowhere/tesseract.exe"},
                logging.getLogger("test"),
            )
        finally:
            module.TESSERACT_AVAILABLE = original


class TestDocumentToPDF(unittest.TestCase):
    """X -> PDF via headless LibreOffice.

    LibreOffice is absent from CI and from most dev machines, so the tests that
    matter here are the ones that do not need it: that a missing binary is
    reported rather than crashed on, that unsupported input is rejected up
    front, and that the command carries the private-profile flag without which
    soffice silently writes nothing. The one test that needs a real
    installation skips itself.
    """

    def setUp(self):
        self.tmp = make_test_dir()
        self.out = os.path.join(self.tmp, "pdfs")
        self.logger = setup_logging(self.tmp)
        self.docx = os.path.join(self.tmp, "report.docx")
        with open(self.docx, "wb") as fh:      # contents never read by these tests
            fh.write(b"not a real docx")

    def tearDown(self):
        close_logger(self.logger)
        shutil.rmtree(self.tmp)

    def _conv(self, **cfg):
        from pdf_to_excel import DocumentToPDFConverter
        base = load_config()
        base.update(cfg)
        return DocumentToPDFConverter(base, self.logger)

    # ── locating the binary ──────────────────────────────────────────────
    def test_missing_libreoffice_is_reported_not_raised(self):
        conv = self._conv(libreoffice_path=None)
        conv._binary = None                    # force "not found"
        r = conv.convert(self.docx, self.out)
        self.assertFalse(r["success"])
        self.assertIn("LibreOffice not found", r["error"])

    def test_configured_path_is_used_when_it_exists(self):
        fake = os.path.join(self.tmp, "soffice.exe")
        open(fake, "w").close()
        self.assertEqual(self._conv(libreoffice_path=fake).binary(), fake)

    def test_wrong_configured_path_falls_back_and_warns(self):
        # A bad path in config is a mistake worth hearing about, not something
        # to paper over by quietly searching elsewhere.
        missing = os.path.join(self.tmp, "nope", "soffice")
        conv = self._conv(libreoffice_path=missing)
        with self.assertLogs(self.logger, level="WARNING") as caught:
            conv.binary()
        self.assertTrue(any("libreoffice_path" in m for m in caught.output))

    # ── input validation, before anything is launched ────────────────────
    def test_unsupported_extension_is_rejected_up_front(self):
        odd = os.path.join(self.tmp, "archive.zip")
        open(odd, "w").close()
        r = self._conv().convert(odd, self.out)
        self.assertFalse(r["success"])
        self.assertIn("Unsupported format", r["error"])

    def test_pdf_input_is_rejected_clearly(self):
        pdf = os.path.join(self.tmp, "already.pdf")
        open(pdf, "w").close()
        r = self._conv().convert(pdf, self.out)
        self.assertFalse(r["success"])
        self.assertEqual(r["error"], "Already a PDF")

    def test_missing_file_is_reported(self):
        r = self._conv().convert(os.path.join(self.tmp, "ghost.docx"), self.out)
        self.assertFalse(r["success"])
        self.assertEqual(r["error"], "File not found")

    # ── the command itself ───────────────────────────────────────────────
    def test_command_isolates_the_user_profile(self):
        # Without -env:UserInstallation a second soffice hands its job to the
        # running instance and exits 0 having written nothing - a silent
        # failure with a success exit code.
        from pdf_to_excel import DocumentToPDFConverter
        cmd = DocumentToPDFConverter._command(
            "soffice", "/in/a.docx", "/out", os.path.join(self.tmp, "profile")
        )
        self.assertTrue(any(c.startswith("-env:UserInstallation=file:") for c in cmd))
        self.assertIn("--headless", cmd)
        self.assertEqual(cmd[cmd.index("--convert-to") + 1], "pdf")
        self.assertEqual(cmd[cmd.index("--outdir") + 1], "/out")
        self.assertEqual(cmd[-1], "/in/a.docx")

    # ── result handling, with soffice stubbed ────────────────────────────
    def test_success_requires_the_pdf_to_actually_exist(self):
        # soffice returns 0 in cases where it wrote nothing, so the output file
        # is what decides success - not the exit code.
        import subprocess as sp
        from pdf_to_excel import DocumentToPDFConverter

        conv = self._conv()
        conv._binary = "soffice"
        completed = sp.CompletedProcess(args=[], returncode=0, stdout="", stderr="")
        with unittest.mock.patch("pdf_to_excel.subprocess.run", return_value=completed):
            r = conv.convert(self.docx, self.out)
        self.assertFalse(r["success"], "exit 0 with no output file must not pass")
        self.assertIn("produced no PDF", r["error"])

    def test_success_when_the_pdf_is_written(self):
        import subprocess as sp
        conv = self._conv()
        conv._binary = "soffice"
        expected = os.path.join(self.out, "report.pdf")

        def fake_run(cmd, **kw):
            # Mimic soffice: write <stem>.pdf into whatever --outdir names,
            # which is a staging directory, not the caller's output folder.
            outdir = cmd[cmd.index("--outdir") + 1]
            os.makedirs(outdir, exist_ok=True)
            with open(os.path.join(outdir, Path(cmd[-1]).stem + ".pdf"), "wb") as fh:
                fh.write(b"%PDF-1.4\n")
            return sp.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

        with unittest.mock.patch("pdf_to_excel.subprocess.run", side_effect=fake_run):
            r = conv.convert(self.docx, self.out)
        self.assertTrue(r["success"], r["error"])
        self.assertEqual(r["output"], expected)
        self.assertTrue(os.path.isfile(expected))

    def test_timeout_is_reported(self):
        import subprocess as sp
        conv = self._conv()
        conv._binary = "soffice"
        with unittest.mock.patch(
            "pdf_to_excel.subprocess.run",
            side_effect=sp.TimeoutExpired(cmd="soffice", timeout=5),
        ):
            r = conv.convert(self.docx, self.out, timeout=5)
        self.assertFalse(r["success"])
        self.assertIn("timed out", r["error"])

    def test_same_stem_in_subfolders_does_not_overwrite(self):
        # The folder walk is recursive, so two documents can share a stem.
        # Letting them share the output name meant the second silently
        # overwrote the first while both were reported converted.
        import subprocess as sp
        for sub, ext in (("jan", ".txt"), ("feb", ".html")):
            folder = os.path.join(self.tmp, "docs", sub)
            os.makedirs(folder, exist_ok=True)
            with open(os.path.join(folder, "report" + ext), "w") as fh:
                fh.write(sub)

        conv = self._conv()
        conv._binary = "soffice"

        def fake_run(cmd, **kw):
            # Mimic soffice: write <stem>.pdf into whatever --outdir says.
            outdir = cmd[cmd.index("--outdir") + 1]
            stem = Path(cmd[-1]).stem
            os.makedirs(outdir, exist_ok=True)
            with open(os.path.join(outdir, stem + ".pdf"), "wb") as fh:
                fh.write(b"%PDF-1.4\n")
            return sp.CompletedProcess(args=cmd, returncode=0, stdout="", stderr="")

        with unittest.mock.patch("pdf_to_excel.subprocess.run", side_effect=fake_run):
            results = conv.convert_folder(os.path.join(self.tmp, "docs"), self.out)

        self.assertEqual(len(results), 2)
        self.assertTrue(all(r["success"] for r in results), results)
        outputs = {r["output"] for r in results}
        self.assertEqual(len(outputs), 2, "both documents wrote to one path")
        for path in outputs:
            self.assertTrue(os.path.isfile(path), f"missing: {path}")

    # ── live check, skipped unless LibreOffice is actually installed ─────
    def test_real_conversion_if_libreoffice_is_installed(self):
        """Drive a real soffice, and check the PDF is readable - not just present.

        A file of the right name proves very little: soffice can emit an empty
        or unopenable document. This asserts the output parses and the source
        text survived the round trip.
        """
        from pdf_to_excel import libreoffice_present
        if not libreoffice_present(load_config()):
            self.skipTest("LibreOffice not installed on this machine")

        import pdfplumber

        # A text document (Writer) and a spreadsheet (Calc): two different
        # LibreOffice filters, so this covers more than one import path.
        txt = os.path.join(self.tmp, "sample.txt")
        with open(txt, "w", encoding="utf-8") as fh:
            fh.write("Supplier order report\nMARKER-TXT\n")

        from openpyxl import Workbook
        wb = Workbook()
        wb.active.append(["S. No.", "Customer Name", "Amount"])
        wb.active.append([1, "MARKER-XLSX", "413.25"])
        # LibreOffice renders a spreadsheet at its stored column widths and
        # clips anything wider - at the default width "MARKER-XLSX" comes out
        # as "MARKER-XL". Widening the column is what a real sheet would carry,
        # and keeps this test measuring conversion rather than truncation.
        wb.active.column_dimensions["B"].width = 24
        xlsx = os.path.join(self.tmp, "sheet.xlsx")
        wb.save(xlsx)

        conv = self._conv()
        for src, marker in ((txt, "MARKER-TXT"), (xlsx, "MARKER-XLSX")):
            with self.subTest(source=os.path.basename(src)):
                r = conv.convert(src, self.out)
                self.assertTrue(r["success"], r["error"])
                self.assertTrue(os.path.isfile(r["output"]))

                with open(r["output"], "rb") as fh:
                    self.assertEqual(fh.read(5), b"%PDF-")

                with pdfplumber.open(r["output"]) as pdf:
                    self.assertGreaterEqual(len(pdf.pages), 1)
                    text = " ".join(
                        (p.extract_text() or "") for p in pdf.pages
                    )
                self.assertIn(marker, text,
                              "source text did not survive into the PDF")


class TestLogging(unittest.TestCase):
    def setUp(self):
        self.tmp = make_test_dir()

    def tearDown(self):
        close_logger(logging.getLogger("pdf_to_excel"))
        shutil.rmtree(self.tmp)

    def test_logger_created(self):
        logger = setup_logging(self.tmp)
        self.assertIsNotNone(logger)

    def test_log_file_created(self):
        setup_logging(self.tmp)
        logs = list(Path(self.tmp).glob("*.log"))
        self.assertEqual(len(logs), 1)


class TestExcelBuilder(unittest.TestCase):
    def setUp(self):
        self.tmp = make_test_dir()
        self.cfg = load_config()
        self.logger = logging.getLogger("test")

    def tearDown(self):
        shutil.rmtree(self.tmp)

    def _make_df(self):
        return pd.DataFrame({
            "Product": ["Widget A", "Widget B", "Gadget X"],
            "Units": [100, 200, 300],
            "Revenue": [1000.0, 2000.0, 3000.0],
        })

    def test_add_and_save_dataframe(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        df = self._make_df()
        builder.add_dataframe(df, "Sales")
        out = os.path.join(self.tmp, "test_output.xlsx")
        success = builder.save(out)
        self.assertTrue(success)
        self.assertTrue(os.path.exists(out))

    def test_output_has_correct_headers(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        df = self._make_df()
        builder.add_dataframe(df, "Sales")
        out = os.path.join(self.tmp, "test_output.xlsx")
        builder.save(out)

        wb = load_workbook(out)
        ws = wb["Sales"]
        headers = [ws.cell(1, c).value for c in range(1, 4)]
        self.assertEqual(headers, ["Product", "Units", "Revenue"])

    def test_output_has_correct_data(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        df = self._make_df()
        builder.add_dataframe(df, "Sales")
        out = os.path.join(self.tmp, "test_output.xlsx")
        builder.save(out)

        wb = load_workbook(out)
        ws = wb["Sales"]
        self.assertEqual(ws.cell(2, 1).value, "Widget A")
        self.assertEqual(ws.cell(3, 2).value, 200)

    def test_metadata_sheet_added(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        self.cfg["add_metadata_sheet"] = True
        builder.add_dataframe(self._make_df(), "Data")
        meta = {"Source": "test.pdf", "Pages": 5}
        builder.add_metadata_sheet(meta)
        out = os.path.join(self.tmp, "meta_test.xlsx")
        builder.save(out)

        wb = load_workbook(out)
        self.assertIn("_Metadata", wb.sheetnames)

    def test_unique_sheet_names(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        df = self._make_df()
        for _ in range(5):
            builder.add_dataframe(df, "Sheet")
        out = os.path.join(self.tmp, "unique_sheets.xlsx")
        builder.save(out)

        wb = load_workbook(out)
        names = wb.sheetnames
        self.assertEqual(len(names), len(set(names)))

    def test_duplicate_column_names_do_not_crash(self):
        # Promoting a header row can produce repeated names. Sizing columns by
        # label then returned a DataFrame instead of a Series and raised, which
        # failed the entire file rather than one column's width.
        df = pd.DataFrame([["a", "b"], ["c", "d"]], columns=["Col", "Col"])
        builder = ExcelBuilder(self.cfg, self.logger)
        builder.add_dataframe(df, "Dupes")
        out = os.path.join(self.tmp, "dupes.xlsx")
        self.assertTrue(builder.save(out))

        ws = load_workbook(out)["Dupes"]
        self.assertEqual([c.value for c in ws[1]], ["Col", "Col"])

    def test_serial_number_is_written_as_a_number(self):
        # Serial numbers stayed text, so Excel flagged the column and sorted
        # 10 before 2. Identifiers around them must still stay text.
        builder = ExcelBuilder(self.cfg, self.logger)
        self.assertEqual(builder._coerce_value("2", "S. No."), 2)
        self.assertIsInstance(builder._coerce_value("2", "S. No."), int)
        self.assertEqual(builder._coerce_value("00123456", "Order No"), "00123456")
        self.assertEqual(builder._coerce_value("$413.25", "Amount"), 413.25)
        # A non-numeric serial is left alone rather than forced.
        self.assertEqual(builder._coerce_value("n/a", "S. No."), "n/a")

    def test_empty_dataframe_skipped(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        builder.add_dataframe(pd.DataFrame(), "Empty")
        # Should have no sheets
        self.assertEqual(len(builder.wb.sheetnames), 0)

    def test_text_sheet(self):
        builder = ExcelBuilder(self.cfg, self.logger)
        builder.add_text_sheet("Line 1\nLine 2\nLine 3", "TextPage")
        out = os.path.join(self.tmp, "text_sheet.xlsx")
        builder.save(out)
        wb = load_workbook(out)
        ws = wb["TextPage"]
        self.assertEqual(ws.cell(1, 1).value, "Line 1")
        self.assertEqual(ws.cell(3, 1).value, "Line 3")


class TestTableCleaning(unittest.TestCase):
    def setUp(self):
        self.cfg = load_config()
        self.logger = logging.getLogger("test")
        self.extractor = TableExtractor(self.cfg, self.logger)

    def test_clean_removes_empty_rows(self):
        import numpy as np
        df = pd.DataFrame([
            ["A", "1"], [None, None], ["B", "2"]
        ])
        cleaned = self.extractor._clean_dataframe(df, promote_header=False)
        self.assertEqual(len(cleaned), 2)

    def test_clean_promotes_header(self):
        df = pd.DataFrame([
            ["Name", "Value"],
            ["Alice", "100"],
            ["Bob", "200"]
        ])
        cleaned = self.extractor._clean_dataframe(df, promote_header=True)
        self.assertIn("Name", cleaned.columns)
        self.assertEqual(len(cleaned), 2)

    def test_clean_empty_df(self):
        df = pd.DataFrame()
        cleaned = self.extractor._clean_dataframe(df)
        self.assertTrue(cleaned.empty)

    def test_clean_merges_split_report_rows(self):
        df = pd.DataFrame([
            ["S. No.", "Customer Name", "Item", "Product Type", "Certificate No", "Order No", "Amount", "Manufacturer", "Est Shipping Date"],
            ["1", "Alex Morgan", "2.14 Round J VS1", "Diamond", "IGI", "A1B2C3D4", 413.25, "Guild and Facet", "27 Apr 2026"],
            [None, None, None, None, "LG_100000001", None, None, None, None],
            ["8", "Jordan Avery", "2.20 Emerald J", "Diamond", "IGI 900000001", "E5F6A7B8", 4722.30, "Guild and Facet", "24 Apr 2026"],
            [None, None, "VVS2", None, None, None, None, None, None],
        ])
        cleaned = self.extractor._clean_dataframe(df, promote_header=True)
        self.assertEqual(len(cleaned), 2)
        self.assertEqual(cleaned.loc[0, "Certificate No"], "IGI LG_100000001")
        self.assertEqual(cleaned.loc[1, "Item"], "2.20 Emerald J VVS2")


@unittest.skipUnless(REPORTLAB_AVAILABLE, "reportlab required for PDF generation tests")
class TestPDFProcessor(unittest.TestCase):
    def setUp(self):
        self.tmp = make_test_dir()
        self.cfg = load_config()
        self.cfg["output_dir"] = self.tmp
        self.cfg["log_dir"] = self.tmp
        self.cfg["add_metadata_sheet"] = True
        self.logger = setup_logging(self.tmp)
        self.processor = PDFProcessor(self.cfg, self.logger)

    def tearDown(self):
        close_logger(self.logger)
        shutil.rmtree(self.tmp)

    def test_single_table_pdf(self):
        pdf = os.path.join(self.tmp, "table_test.pdf")
        create_test_pdf_with_table(pdf)
        out = os.path.join(self.tmp, "table_test.xlsx")
        result = self.processor.process(pdf, out)
        self.assertTrue(result["success"], f"Failed: {result['error']}")
        self.assertTrue(os.path.exists(out))
        self.assertGreater(result["pages"], 0)

    def test_multi_section_pdf(self):
        pdf = os.path.join(self.tmp, "multi_test.pdf")
        create_multi_table_pdf(pdf)
        out = os.path.join(self.tmp, "multi_test.xlsx")
        result = self.processor.process(pdf, out)
        self.assertTrue(result["success"], f"Failed: {result['error']}")
        wb = load_workbook(out)
        self.assertGreater(len(wb.sheetnames), 0)

    def test_text_only_pdf(self):
        pdf = os.path.join(self.tmp, "text_test.pdf")
        create_text_only_pdf(pdf)
        out = os.path.join(self.tmp, "text_test.xlsx")
        result = self.processor.process(pdf, out)
        self.assertTrue(result["success"], f"Failed: {result['error']}")

    def test_metadata_sheet_present(self):
        pdf = os.path.join(self.tmp, "meta_check.pdf")
        create_test_pdf_with_table(pdf)
        out = os.path.join(self.tmp, "meta_check.xlsx")
        self.processor.process(pdf, out)
        wb = load_workbook(out)
        self.assertIn("_Metadata", wb.sheetnames)

    def test_nonexistent_file(self):
        result = self.processor.process("/no/such/file.pdf", "/tmp/out.xlsx")
        self.assertFalse(result["success"])
        self.assertIsNotNone(result["error"])

    def test_empty_file(self):
        empty = os.path.join(self.tmp, "empty.pdf")
        open(empty, "w").close()
        out = os.path.join(self.tmp, "empty_out.xlsx")
        result = self.processor.process(empty, out)
        self.assertFalse(result["success"])

    def test_text_pages_combined_into_one_sheet(self):
        pdf = os.path.join(self.tmp, "text_pages.pdf")
        with open(pdf, "wb") as f:
            f.write(b"%PDF-1.4 dummy")
        out = os.path.join(self.tmp, "text_pages.xlsx")

        self.processor.classifier.classify = lambda _: {
            "page_count": 2,
            "type": "text",
            "text_pages": [0, 1],
            "scanned_pages": [],
        }
        self.processor.table_extractor.extract = lambda *_: []
        self.processor.text_extractor.extract = lambda *_: [
            {
                "page": 1,
                "text": "page1",
                "dataframe": pd.DataFrame([{
                    "S. No.": "1",
                    "Customer Name": "Alex Morgan",
                    "Item": "2.14 Round J VS1",
                }]),
            },
            {
                "page": 2,
                "text": "page2",
                "dataframe": pd.DataFrame([{
                    "S. No.": "2",
                    "Customer Name": "Daniel Hesse",
                    "Item": "3.41 Radiant F VVS2",
                }]),
            },
        ]

        result = self.processor.process(pdf, out)
        self.assertTrue(result["success"], f"Failed: {result['error']}")
        wb = load_workbook(out)
        self.assertIn("Text_Data", wb.sheetnames)
        self.assertNotIn("P1_Text", wb.sheetnames)
        ws = wb["Text_Data"]
        # Serial numbers are written as numbers, so Excel sorts the column
        # correctly and stops flagging "number stored as text".
        self.assertEqual(ws.cell(4, 1).value, 1)
        self.assertEqual(ws.cell(5, 3).value, "3.41 Radiant F VVS2")

    def test_combine_extracted_tables_deduplicates(self):
        df1 = pd.DataFrame([{
            "S. No.": "25",
            "Customer Name": "Joel Chavez Flores",
            "Item": "1.00 Princess D VVS2",
        }])
        df2 = pd.DataFrame([{
            "S. No.": "25",
            "Customer Name": "Joel Chavez Flores",
            "Item": "1.00 Princess D VVS2",
        }, {
            "S. No.": "26",
            "Customer Name": "Paul Hildebrand",
            "Item": "1.57 Round H VS2",
        }])
        combined = self.processor._combine_extracted_tables([
            {"dataframe": df1},
            {"dataframe": df2},
        ])
        self.assertEqual(len(combined), 2)


@unittest.skipUnless(REPORTLAB_AVAILABLE, "reportlab required for batch tests")
class TestBatchProcessor(unittest.TestCase):
    def setUp(self):
        self.tmp = make_test_dir()
        self.input_dir = os.path.join(self.tmp, "input")
        self.output_dir = os.path.join(self.tmp, "output")
        os.makedirs(self.input_dir)
        self.cfg = load_config()
        self.cfg["log_dir"] = self.tmp
        self.logger = setup_logging(self.tmp)

    def tearDown(self):
        close_logger(self.logger)
        shutil.rmtree(self.tmp)

    def _make_pdfs(self, n: int) -> None:
        for i in range(1, n + 1):
            pdf = os.path.join(self.input_dir, f"report_{i}.pdf")
            create_test_pdf_with_table(pdf)

    def test_batch_all_succeed(self):
        self._make_pdfs(3)
        batch = BatchProcessor(self.cfg, self.logger)
        results = batch.process_folder(self.input_dir, self.output_dir)
        self.assertEqual(len(results), 3)
        self.assertTrue(all(r["success"] for r in results), results)

    def test_batch_summary_report_created(self):
        self._make_pdfs(2)
        batch = BatchProcessor(self.cfg, self.logger)
        batch.process_folder(self.input_dir, self.output_dir)
        reports = list(Path(self.output_dir).glob("_batch_summary_*.xlsx"))
        self.assertEqual(len(reports), 1)

    def test_batch_empty_folder(self):
        batch = BatchProcessor(self.cfg, self.logger)
        results = batch.process_folder(self.input_dir, self.output_dir)
        self.assertEqual(len(results), 0)

    def test_same_stem_in_subfolders_does_not_overwrite(self):
        # Batch mode recurses subdirectories, so two different PDFs can share a
        # stem. Naming the output from the stem alone made the second overwrite
        # the first while the summary reported both as succeeded - one whole
        # converted workbook lost with no indication anything had happened.
        for sub in ("jan", "feb"):
            folder = os.path.join(self.input_dir, sub)
            os.makedirs(folder)
            create_test_pdf_with_table(os.path.join(folder, "orders.pdf"))

        batch = BatchProcessor(self.cfg, self.logger)
        results = batch.process_folder(self.input_dir, self.output_dir)

        self.assertEqual(len(results), 2)
        self.assertTrue(all(r["success"] for r in results), results)

        outputs = {r["output"] for r in results}
        self.assertEqual(len(outputs), 2, "both PDFs were written to one path")
        for path in outputs:
            self.assertTrue(os.path.exists(path), f"missing output: {path}")

        workbooks = [
            p for p in Path(self.output_dir).glob("*.xlsx")
            if not p.name.startswith("_batch_summary")
        ]
        self.assertEqual(len(workbooks), 2)

    def test_processing_order_is_deterministic(self):
        # Deduplicating through a set discarded the sort, leaving the order to
        # set iteration. Files are processed in sorted order.
        self._make_pdfs(5)
        batch = BatchProcessor(self.cfg, self.logger)
        names = [
            os.path.basename(r["input"])
            for r in batch.process_folder(self.input_dir, self.output_dir)
        ]
        self.assertEqual(names, sorted(names))

    def test_batch_skips_corrupt_files(self):
        self._make_pdfs(2)
        # Add a corrupt PDF
        bad = os.path.join(self.input_dir, "corrupt.pdf")
        with open(bad, "w") as f:
            f.write("NOT A PDF")
        batch = BatchProcessor(self.cfg, self.logger)
        results = batch.process_folder(self.input_dir, self.output_dir)
        # Processing should complete; corrupt file will fail gracefully
        self.assertEqual(len(results), 3)
        ok = sum(1 for r in results if r["success"])
        self.assertGreaterEqual(ok, 2)


class TestWebApp(unittest.TestCase):
    """app.py had no automated coverage at all.

    This does not exercise a conversion - that needs an uploaded file - but it
    does prove the module imports, the token/CSS block renders, and the first
    viewport reaches its stop point without raising. Streamlit re-runs the whole
    script on every interaction, so a module-level error there breaks every
    screen at once.
    """

    def test_first_viewport_renders_without_exception(self):
        try:
            from streamlit.testing.v1 import AppTest
        except ImportError:  # older streamlit without the testing harness
            self.skipTest("streamlit.testing.v1 unavailable")

        app_path = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "app.py"
        )
        at = AppTest.from_file(app_path, default_timeout=60).run()

        self.assertFalse(
            at.exception,
            f"app.py raised on first render: {[e.value for e in at.exception]}",
        )
        # The uploader is the entry point; without it there is no way in.
        self.assertTrue(at.markdown, "app rendered no markdown at all")


class TestMixedTableAndReport(unittest.TestCase):
    """A ruled table on one page must not hide a report on another."""

    def setUp(self):
        self.tmp = make_test_dir()
        self.cfg = load_config()
        self.cfg["log_dir"] = self.tmp
        self.logger = setup_logging(self.tmp)
        self.pdf = os.path.join(self.tmp, "mixed.pdf")
        create_table_then_report_pdf(self.pdf)

    def tearDown(self):
        close_logger(self.logger)
        shutil.rmtree(self.tmp)

    def test_report_page_is_read_despite_a_table_on_another_page(self):
        processor = PDFProcessor(self.cfg, self.logger)
        info = processor.classifier.classify(self.pdf)
        datasets, _text = processor.extract_datasets(self.pdf, info)

        # The report's rows must appear somewhere in the extracted datasets.
        report_frames = [
            ds["dataframe"] for ds in datasets
            if "Order No" in list(ds["dataframe"].columns)
        ]
        self.assertTrue(
            report_frames,
            f"report page not read; datasets were {[d['name'] for d in datasets]}"
        )
        orders = set()
        for frame in report_frames:
            orders.update(frame["Order No"].astype(str))
        self.assertIn("A1B2C3D4", orders)
        self.assertIn("E5F6A7B8", orders)

    def test_both_pages_reach_the_workbook(self):
        out = os.path.join(self.tmp, "mixed.xlsx")
        result = PDFProcessor(self.cfg, self.logger).process(self.pdf, out)
        self.assertTrue(result["success"], result)
        self.assertGreaterEqual(result["rows_extracted"], 2)


class TestCsvExport(unittest.TestCase):
    """CSV export must see everything the Excel path sees.

    It used to call the table engines directly, so it covered only documents
    with a ruled grid. Order reports carry no ruling and are read further down
    the cascade by the column parser, which meant exporting one as CSV wrote no
    files at all and said "no tables found" - on the document the tool exists
    for.
    """

    def setUp(self):
        self.tmp = make_test_dir()
        self.out = os.path.join(self.tmp, "csv")
        self.cfg = load_config()
        self.cfg["log_dir"] = self.tmp
        self.logger = setup_logging(self.tmp)

    def tearDown(self):
        close_logger(self.logger)
        shutil.rmtree(self.tmp)

    def test_order_report_exports_its_rows(self):
        pdf = os.path.join(self.tmp, "report.pdf")
        create_order_report_pdf(pdf)

        export = export_to_csv(pdf, self.out, self.cfg, self.logger)

        self.assertEqual(len(export["files"]), 1, export["files"])
        self.assertEqual(export["unparsed_rows"], 0)

        df = pd.read_csv(export["files"][0], dtype=str)
        self.assertEqual(list(df.columns), REPORT_COLUMNS)
        self.assertEqual(len(df), 2)
        self.assertEqual(df.loc[0, "Order No"], "A1B2C3D4")
        self.assertEqual(df.loc[1, "Amount"], "$4722.30")

    def test_ruled_table_still_exports(self):
        # The table path must keep working - the fix shares the cascade, it
        # does not replace one branch with another.
        pdf = os.path.join(self.tmp, "table.pdf")
        create_test_pdf_with_table(pdf)

        export = export_to_csv(pdf, self.out, self.cfg, self.logger)

        self.assertTrue(export["files"])
        self.assertTrue(all(p.endswith(".csv") for p in export["files"]))

    def test_text_only_pdf_exports_instead_of_nothing(self):
        pdf = os.path.join(self.tmp, "text.pdf")
        create_text_only_pdf(pdf)

        export = export_to_csv(pdf, self.out, self.cfg, self.logger)

        self.assertTrue(export["files"], "text-only PDF exported no CSV")

    def test_csv_filenames_match_sheet_names(self):
        pdf = os.path.join(self.tmp, "report.pdf")
        create_order_report_pdf(pdf)

        export = export_to_csv(pdf, self.out, self.cfg, self.logger)
        names = [os.path.basename(p) for p in export["files"]]
        self.assertEqual(names, ["report_Text_Data.csv"])


class TestTextExtractor(unittest.TestCase):
    def setUp(self):
        self.logger = logging.getLogger("test")
        self.extractor = TextExtractor(self.logger)

    def test_text_to_dataframe_can_skip_report_parsing(self):
        # Used by the caller for a page the column parser already read and
        # already reported on, so its exclusions are not counted twice.
        self.extractor.dropped_rows = []
        df = self.extractor._text_to_dataframe(
            "\n".join(self.REPORT_FIXTURE), parse_report=False
        )
        self.assertNotIn("Order No", list(df.columns))
        self.assertEqual(self.extractor.dropped_rows, [])

    def test_parse_order_report_text(self):
        text = "\n".join([
            "Invoice #GF2026042110",
            "Created Date - 21-Apr-2026",
            "95 Items - $94222.52",
            "S. No. Customer Name Item Product Type Certificate No Order No Amount Manufacturer Est Shipping Date",
            "1 Alex Morgan 2.14 Round J VS1 Diamond IGI A1B2C3D4 $413.25 Guild and Facet 27 Apr 2026",
            "LG_100000001",
            "8 Jordan Avery 2.20 Emerald J Diamond IGI 900000001 E5F6A7B8 $4722.30 Guild and Facet 24 Apr 2026",
            "VVS2",
        ])
        df = self.extractor._text_to_dataframe(text)
        self.assertEqual(list(df.columns), [
            "S. No.", "Customer Name", "Item", "Product Type",
            "Certificate No", "Order No", "Amount",
            "Manufacturer", "Est Shipping Date"
        ])
        self.assertEqual(df.loc[0, "Item"], "2.14 Round J VS1")
        self.assertEqual(df.loc[0, "Certificate No"], "IGI LG_100000001")
        self.assertEqual(df.loc[1, "Item"], "2.20 Emerald J VVS2")

    # ── Footer / trailing-content handling ────────────────────────────────
    REPORT_FIXTURE = [
        "Invoice #GF2026042110",
        "Created Date - 21-Apr-2026",
        "95 Items - $94222.52",
        "S. No. Customer Name Item Product Type Certificate No Order No "
        "Amount Manufacturer Est Shipping Date",
        "1 Alex Morgan 2.14 Round J VS1 Diamond IGI A1B2C3D4 $413.25 "
        "Guild and Facet 27 Apr 2026",
        "LG_100000001",
        "8 Jordan Avery 2.20 Emerald J Diamond IGI 900000001 E5F6A7B8 "
        "$4722.30 Guild and Facet 24 Apr 2026",
        "VVS2",
    ]

    def test_footer_does_not_contaminate_last_row(self):
        lines = self.REPORT_FIXTURE + [
            "Total: $94,222.52",
            "Thank you for your business",
            "Page 1 of 1",
            "Please remit payment within 30 days",
        ]
        df = self.extractor._parse_order_report(lines)
        self.assertEqual(len(df), 2)
        # Previously every trailing line was appended to the preceding row.
        self.assertEqual(df.loc[1, "Item"], "2.20 Emerald J VVS2")
        for value in df.loc[1].tolist():
            self.assertNotIn("Total", str(value))
            self.assertNotIn("Thank", str(value))
            self.assertNotIn("remit", str(value))

    def test_digit_leading_footer_is_not_a_data_row(self):
        # "95 Items - $94222.52" starts with digits and would otherwise be
        # picked up as a row, then reported as an unparsed one.
        self.extractor.dropped_rows = []
        df = self.extractor._parse_order_report(
            self.REPORT_FIXTURE + ["95 Items - $94222.52"]
        )
        self.assertEqual(len(df), 2)
        self.assertEqual(self.extractor.dropped_rows, [])

    def test_unparseable_data_row_is_recorded(self):
        self.extractor.dropped_rows = []
        df = self.extractor._parse_order_report(
            self.REPORT_FIXTURE
            + ["9 Casey Lin 1.01 Round F VS2 Sapphire IGI 111222333 AB12CD34 "
               "$999.00 Guild and Facet 30 Apr 2026"]
        )
        self.assertEqual(len(df), 2)
        self.assertEqual(len(self.extractor.dropped_rows), 1)

    def test_junk_between_rows_does_not_stop_parsing(self):
        lines = self.REPORT_FIXTURE[:6] + [
            "-- continued --",
            "8 Jordan Avery 2.20 Emerald J Diamond IGI 900000001 "
            "E5F6A7B8 $4722.30 Guild and Facet 24 Apr 2026",
            "VVS2",
        ]
        df = self.extractor._parse_order_report(lines)
        self.assertEqual(len(df), 2)
        self.assertEqual(df.loc[1, "Item"], "2.20 Emerald J VVS2")

    def test_cut_grade_continuation_merges_certificate(self):
        # Real reports wrap as "Excellent LG_100000002"; the cut grade must not
        # cause the whole line - and with it the certificate id - to be skipped.
        lines = [
            "S. No. Customer Name Item Product Type Certificate No Order No "
            "Amount Manufacturer Est Shipping Date",
            "10 Morgan Reid 1.00 Princess E VS2 Diamond IGI C9D0E1F2 $99.00 "
            "USNY 27 Jul 2026",
            "Excellent LG_100000002",
        ]
        df = self.extractor._parse_order_report(lines)
        self.assertEqual(len(df), 1)
        self.assertEqual(df.loc[0, "Certificate No"], "IGI LG_100000002")

    def test_continuation_line_classification(self):
        for line in ["VVS2", "LG_100000001", "IGI 900000001",
                     "2.20 Emerald J", "J", "Excellent LG_100000002",
                     "Very Good LG_123456789"]:
            self.assertTrue(
                self.extractor._is_continuation_line(line), f"expected continuation: {line}"
            )
        for line in ["Total: $94,222.52", "Page 1 of 3",
                     "Please remit payment to the address above",
                     "Example Diamonds NV, Antwerp"]:
            self.assertFalse(
                self.extractor._is_continuation_line(line), f"expected footer: {line}"
            )

    def test_text_to_dataframe_plain_text(self):
        text = "Apple\nBanana\nCherry"
        df = self.extractor._text_to_dataframe(text)
        self.assertFalse(df.empty)

    def test_text_to_dataframe_tab_delimited(self):
        text = "Name\tAge\tCity\nAlice\t30\tNYC\nBob\t25\tLA"
        df = self.extractor._text_to_dataframe(text)
        if "Name" in df.columns:
            self.assertIn("Age", df.columns)

    def test_text_to_dataframe_empty(self):
        df = self.extractor._text_to_dataframe("")
        self.assertTrue(df.empty)


# ─────────────────────────────────────────────────────────────────────────────
# REPORT COLUMN PARSER  (positional geometry, not text patterns)
# ─────────────────────────────────────────────────────────────────────────────

def _w(text, x0, x1, top):
    """A pdfplumber-style word: only top / x0 / x1 are read by the parser."""
    return {"text": text, "x0": x0, "x1": x1, "top": top, "bottom": top + 10}


class _FakePage:
    def __init__(self, words):
        self._words = words

    def extract_words(self):
        return self._words


class _FakePdf:
    def __init__(self, page):
        self.pages = [page]


# Header labels (centred over their columns) and a single valid data row,
# left-aligned under those columns with a clear empty strip between each pair.
_HEADER_WORDS = [
    _w("S.", 50, 58, 100), _w("No.", 60, 72, 100),
    _w("Customer", 90, 140, 100), _w("Name", 142, 175, 100),
    _w("Item", 200, 228, 100),
    _w("Product", 330, 375, 100), _w("Type", 377, 405, 100),
    _w("Certificate", 410, 470, 100), _w("No", 472, 488, 100),
    _w("Order", 500, 535, 100), _w("No", 537, 553, 100),
    _w("Amount", 590, 635, 100),
    _w("Manufacturer", 660, 725, 100),
    _w("Est", 740, 758, 100), _w("Shipping", 760, 805, 100), _w("Date", 807, 835, 100),
]

_DATA_ROW1 = [
    _w("1", 50, 56, 120),
    _w("John", 90, 118, 120), _w("Smith", 120, 150, 120),
    _w("Round", 200, 235, 120),
    _w("Diamond", 330, 375, 120),
    _w("IGI", 410, 430, 120),
    _w("ORD123456", 500, 560, 120),
    _w("$1,234.56", 590, 640, 120),
    _w("ACME", 660, 695, 120),
    _w("15", 740, 752, 120), _w("Jan", 754, 772, 120), _w("2025", 774, 800, 120),
]


class TestReportColumnParser(unittest.TestCase):
    """The positional parser must never silently glue a non-row line onto the
    row above it — that would corrupt or absorb a real order (product principle
    2, 'never fail silently')."""

    def setUp(self):
        self.logger = logging.getLogger("test")
        self.parser = ReportColumnParser(self.logger)
        self.extractor = TextExtractor(self.logger)
        self.extractor.column_parser = self.parser

    def _page(self, *extra_words):
        words = list(_HEADER_WORDS) + list(_DATA_ROW1)
        for group in extra_words:
            words += group
        return _FakePage(words)

    def test_misread_row_is_recorded_not_merged(self):
        # A second order whose serial did not extract: no digit in the S. No.
        # column, but a real Amount and shipping date. It must be reported as
        # unparsed, never merged into the row above.
        misread = [
            _w("Jane", 90, 118, 140), _w("Doe", 120, 145, 140),
            _w("Oval", 200, 232, 140),
            _w("Diamond", 330, 375, 140),
            _w("GIA", 410, 432, 140),
            _w("ORD999999", 500, 560, 140),
            _w("$500.00", 590, 632, 140),
            _w("ACME", 660, 695, 140),
            _w("20", 740, 752, 140), _w("Feb", 754, 772, 140), _w("2025", 774, 800, 140),
        ]
        df = self.extractor._report_by_columns(_FakePdf(self._page(misread)), 0)

        # The kept row is intact — not lengthened by the misread line.
        self.assertEqual(len(df), 1)
        self.assertEqual(df.loc[0, "Amount"], "$1,234.56")
        self.assertEqual(df.loc[0, "Customer Name"], "John Smith")
        self.assertEqual(df.loc[0, "Item"], "Round")
        # The misread order is surfaced, not silently dropped.
        self.assertEqual(len(self.extractor.dropped_rows), 1)
        self.assertIn("ORD999999", self.extractor.dropped_rows[0])

    def test_descriptive_wrap_still_merges(self):
        # Wrapped descriptive cells (Item, Certificate No) carry no Amount or
        # date, so they must merge exactly as before — the fix must not regress
        # the parser's central capability.
        item_tail = [_w("Brilliant", 200, 245, 140)]
        cert_tail = [_w("LG_100000003", 410, 478, 160)]
        rows = self.parser.parse_page(self._page(item_tail, cert_tail))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Item"], "Round Brilliant")
        self.assertEqual(rows[0]["Certificate No"], "IGI LG_100000003")
        self.assertEqual(self.parser.unparsed_lines, [])

    def test_leading_stray_line_is_recorded(self):
        # A non-row, non-footer line before the first data row has no row to
        # attach to; it must be recorded rather than vanishing.
        stray = [_w("Reference", 90, 140, 110), _w("ABC", 200, 225, 110)]
        page = _FakePage(list(_HEADER_WORDS) + stray + list(_DATA_ROW1))
        rows = self.parser.parse_page(page)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Item"], "Round")
        self.assertEqual(self.parser.unparsed_lines, ["Reference ABC"])

    def test_footer_does_not_skew_column_calibration(self):
        # Column boundaries are calibrated from the widest empty strip between
        # two header anchors. A footer printed under the leftmost column -
        # "Total: $94,222.52" - puts words squarely in the strip that separates
        # S. No. from Customer Name. Calibrating from it moved the split to the
        # right, so the first name was swallowed into the serial column, the
        # serial stopped being a bare digit, and every row on the page was
        # rejected. The row loop already skipped footers; the geometry did not.
        footer = [_w("Total:", 50, 74, 160), _w("$94,222.52", 74, 140, 160)]
        rows = self.parser.parse_page(self._page(footer))

        self.assertEqual(len(rows), 1, f"footer skewed calibration: {rows}")
        self.assertEqual(rows[0]["S. No."], "1")
        self.assertEqual(rows[0]["Customer Name"], "John Smith")
        self.assertEqual(self.parser.unparsed_lines, [])

    def test_page_with_no_valid_rows_still_reports_every_line(self):
        # A page the parser read, but where nothing survived validation. The
        # method used to return before surfacing the lines it could attach to
        # no row, so a stray line vanished entirely; and returning None sent the
        # caller into the text parser, which reported the same failures again.
        stray = [_w("Reference", 90, 140, 110), _w("ABC", 200, 225, 110)]
        invalid = [                       # a real order, but with no Amount
            _w("2", 50, 56, 140),
            _w("Jane", 90, 118, 140), _w("Doe", 120, 145, 140),
            _w("Oval", 200, 232, 140),
            _w("Diamond", 330, 375, 140),
            _w("GIA", 410, 432, 140),
            _w("ORD999999", 500, 560, 140),
            _w("ACME", 660, 695, 140),
            _w("20", 740, 752, 140), _w("Feb", 754, 772, 140),
            _w("2025", 774, 800, 140),
        ]
        page = _FakePage(list(_HEADER_WORDS) + stray + invalid)

        df = self.extractor._report_by_columns(_FakePdf(page), 0)

        # Empty rather than None: the layout WAS read, so the caller must not
        # re-run the text parser and count these exclusions a second time.
        self.assertIsNotNone(df)
        self.assertTrue(df.empty)

        # Both the failed row and the unattachable line are reported, once each.
        self.assertEqual(len(self.extractor.dropped_rows), 2,
                         self.extractor.dropped_rows)
        joined = " | ".join(self.extractor.dropped_rows)
        self.assertIn("ORD999999", joined)
        self.assertIn("Reference ABC", joined)

    def test_unparsed_lines_do_not_leak_between_pages(self):
        # unparsed_lines was cleared beside the row loop, below several early
        # returns. A page that bailed out before reaching it left the previous
        # page's lines in place for the caller to read.
        stray = [_w("Reference", 90, 140, 110), _w("ABC", 200, 225, 110)]
        self.parser.parse_page(_FakePage(list(_HEADER_WORDS) + stray + list(_DATA_ROW1)))
        self.assertEqual(self.parser.unparsed_lines, ["Reference ABC"])

        # A page with no readable header returns early; nothing must survive.
        self.parser.parse_page(_FakePage([_w("Unrelated", 50, 110, 100)]))
        self.assertEqual(self.parser.unparsed_lines, [])

    def test_recognised_footer_is_neither_merged_nor_recorded(self):
        # A footer the marker recognises is skipped outright: it does not touch
        # the row above and does not raise a spurious unparsed warning.
        footer = [
            _w("Page", 50, 80, 140), _w("1", 82, 90, 140),
            _w("of", 92, 105, 140), _w("1", 107, 113, 140),
        ]
        rows = self.parser.parse_page(self._page(footer))

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["Item"], "Round")
        self.assertEqual(self.parser.unparsed_lines, [])


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 60)
    print("PDF to Excel Converter – Test Suite")
    print("=" * 60)

    if not REPORTLAB_AVAILABLE:
        print(
            "\n[WARN] reportlab not installed. PDF-generation tests will be skipped.\n"
            "       Install with: pip install reportlab\n"
        )

    # Show dependency status before running tests
    from pdf_to_excel import print_dependency_status
    print_dependency_status()
    print()

    unittest.main(verbosity=2)
