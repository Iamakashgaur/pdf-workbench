#!/usr/bin/env python3
"""
PDF to Excel Converter
Handles scanned, native, and mixed PDFs via OCR + layout analysis.
"""

import os
import sys
import json
import logging
import argparse
import warnings
import io
import re
import csv
import shutil
import tempfile
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Optional, Union, List, Dict, Any, Tuple

warnings.filterwarnings("ignore")

# ──────────────── Core ────────────────────────────────────────────────────────────
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────── PDF ────────────────────────────────────────────────────────────
try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import fitz  # PyMuPDF
    PYMUPDF_AVAILABLE = True
except ImportError:
    PYMUPDF_AVAILABLE = False

try:
    import camelot
    CAMELOT_AVAILABLE = True
except ImportError:
    CAMELOT_AVAILABLE = False

try:
    import tabula
    TABULA_AVAILABLE = True
except ImportError:
    TABULA_AVAILABLE = False

# ──────────────── OCR ────────────────────────────────────────────────────────────
try:
    from PIL import Image, ImageEnhance, ImageFilter
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False

try:
    import pytesseract
    TESSERACT_AVAILABLE = True
except ImportError:
    TESSERACT_AVAILABLE = False

# ──────────────── Rich UI ────────────────────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.progress import (
        Progress, SpinnerColumn, TextColumn, BarColumn,
        TaskProgressColumn, TimeElapsedColumn
    )
    from rich.table import Table
    from rich.panel import Panel
    from rich.text import Text
    RICH_AVAILABLE = True
    console = Console()
except ImportError:
    RICH_AVAILABLE = False
    console = None


# ────────────────────────────────────────────────────────────
# CONFIGURATION
# ────────────────────────────────────────────────────────────

DEFAULT_CONFIG = {
    "output_dir": "output",
    "log_dir": "logs",
    "ocr_language": "eng",
    "ocr_dpi": 300,
    "ocr_confidence_threshold": 60,
    "table_extraction_method": "auto",  # auto | pdfplumber | camelot | tabula
    "max_col_width": 50,
    "min_col_width": 8,
    "remove_empty_rows": True,
    "remove_empty_cols": True,
    "add_metadata_sheet": True,
    "freeze_header_row": True,
    "apply_table_style": True,
    "header_color": "1F4E79",
    "alternating_row_color": "D6E4F0",
    "batch_summary_report": True,
    "tesseract_path": None,
    "java_path": None,
    "libreoffice_path": None,
    # Extra report shapes to recognise, tried before the built-in one.
    # See README "Report layouts". None/absent means the built-in only.
    "report_layouts": None,
    "preview_rows": 10,
    "max_sheet_name_length": 31
}


def load_config(config_path: str = "config.json") -> dict:
    cfg = DEFAULT_CONFIG.copy()
    if os.path.exists(config_path):
        try:
            with open(config_path) as f:
                user_cfg = json.load(f)
            cfg.update(user_cfg)
            # config.json is committed. A credential in it would be published
            # by the next `git push`, so say so loudly rather than honouring it
            # quietly.
            leaked = [k for k in user_cfg
                      if "password" in k.lower() or "passwd" in k.lower()]
            if leaked:
                print(
                    f"[WARN] {config_path} contains {leaked}. This file is "
                    f"committed to the repository - a password does not belong "
                    f"in it. Use --password or the {PASSWORD_ENV_VAR} "
                    f"environment variable instead."
                )
        except Exception as e:
            print(f"[WARN] Could not load {config_path}: {e}. Using defaults.")
    return cfg


# ────────────────────────────────────────────────────────────
# LOGGING
# ────────────────────────────────────────────────────────────

def setup_logging(log_dir: str, verbose: bool = False) -> logging.Logger:
    os.makedirs(log_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"pdf_to_excel_{ts}.log")

    level = logging.DEBUG if verbose else logging.INFO
    logger = logging.getLogger("pdf_to_excel")
    logger.setLevel(level)
    logger.propagate = False

    for handler in list(logger.handlers):
        handler.close()
        logger.removeHandler(handler)

    fh = logging.FileHandler(log_file, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s | %(levelname)-8s | %(message)s")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    if not verbose:
        ch = logging.StreamHandler()
        ch.setLevel(logging.WARNING)
        ch.setFormatter(fmt)
        logger.addHandler(ch)

    return logger


# ────────────────────────────────────────────────────────────
# SHARED DOCUMENT HANDLES
# ────────────────────────────────────────────────────────────

_UNOPENED = object()   # handle not requested yet
_OPEN_FAILED = object()  # open was attempted and failed; do not retry


class _PDFHandles:
    """Open a PDF once and reuse the handle across a page loop.

    Extractors used to call pdfplumber.open()/fitz.open() per page, which
    reopened and reparsed the whole document for every page. Callers now build
    one of these per document and pass it down; handles are opened lazily so a
    document is only touched by the libraries that actually get used.

    State is tracked with explicit sentinels rather than truthiness: a
    fitz.Document defines __len__, so a zero-page document is falsy and would
    otherwise be mistaken for "not open" and never closed.
    """

    def __init__(self, pdf_path: str, logger: logging.Logger,
                 password: Optional[str] = None):
        self.pdf_path = pdf_path
        self.logger = logger
        # Empty string is the right default: both libraries treat it as "no
        # password", and it keeps every call site free of None-handling.
        self.password = password or ""
        self._plumber: Any = _UNOPENED
        self._fitz: Any = _UNOPENED

    def plumber(self) -> Optional[Any]:
        if self._plumber is _UNOPENED:
            if not PDFPLUMBER_AVAILABLE:
                self._plumber = _OPEN_FAILED
            else:
                try:
                    self._plumber = pdfplumber.open(
                        self.pdf_path, password=self.password)
                except Exception as e:
                    self.logger.debug(f"pdfplumber open failed for '{self.pdf_path}': {e}")
                    self._plumber = _OPEN_FAILED
        return None if self._plumber is _OPEN_FAILED else self._plumber

    def fitz(self) -> Optional[Any]:
        if self._fitz is _UNOPENED:
            if not PYMUPDF_AVAILABLE:
                self._fitz = _OPEN_FAILED
            else:
                try:
                    doc = fitz.open(self.pdf_path)
                    # An encrypted document opens but yields nothing until it
                    # is authenticated, so a page loop would quietly read an
                    # empty file. Treat a failed unlock as a failed open.
                    if doc.needs_pass and not doc.authenticate(self.password):
                        self.logger.debug(
                            f"PyMuPDF could not unlock '{self.pdf_path}'")
                        doc.close()
                        self._fitz = _OPEN_FAILED
                    else:
                        self._fitz = doc
                except Exception as e:
                    self.logger.debug(f"PyMuPDF open failed for '{self.pdf_path}': {e}")
                    self._fitz = _OPEN_FAILED
        return None if self._fitz is _OPEN_FAILED else self._fitz

    def close(self) -> None:
        for handle in (self._plumber, self._fitz):
            if handle is _UNOPENED or handle is _OPEN_FAILED:
                continue
            try:
                handle.close()
            except Exception:
                pass
        self._plumber = _UNOPENED
        self._fitz = _UNOPENED

    def __enter__(self) -> "_PDFHandles":
        return self

    def __exit__(self, *exc_info: Any) -> None:
        self.close()


#: Runtime-only. A password is deliberately NOT a config.json setting: that
#: file is committed, and a credential in a tracked file is a footgun this tool
#: should not hand anyone. It arrives from --password or the environment.
PASSWORD_ENV_VAR = "PDF_TO_EXCEL_PASSWORD"

ACCESS_OK = "ok"
ACCESS_PASSWORD_REQUIRED = "password_required"
ACCESS_PASSWORD_INCORRECT = "password_incorrect"
ACCESS_UNREADABLE = "unreadable"


def check_pdf_access(pdf_path: str, password: Optional[str] = None) -> str:
    """Say whether a PDF can actually be read, before anything tries.

    Returns one of the ACCESS_* constants. The distinction between "needs a
    password" and "that password is wrong" matters: they are different things
    for the person holding the file, and collapsing them into one failure would
    leave them guessing which.

    Without PyMuPDF this cannot be determined, so it reports ACCESS_OK and lets
    the normal path fail in its own way rather than inventing a verdict.
    """
    if not PYMUPDF_AVAILABLE:
        return ACCESS_OK

    doc = None
    try:
        doc = fitz.open(pdf_path)
        if not doc.needs_pass:
            return ACCESS_OK
        if password and doc.authenticate(password):
            return ACCESS_OK
        return ACCESS_PASSWORD_INCORRECT if password else ACCESS_PASSWORD_REQUIRED
    except Exception:
        return ACCESS_UNREADABLE
    finally:
        if doc is not None:
            try:
                doc.close()
            except Exception:
                pass


def _flush_page_cache(page: Any) -> None:
    """Release a pdfplumber page's cached objects.

    Reusing one open document across a long page loop would otherwise grow
    memory monotonically, since the previous code freed it by closing the file.
    """
    flush = getattr(page, "flush_cache", None)
    if callable(flush):
        try:
            flush()
        except Exception:
            pass


# ────────────────────────────────────────────────────────────
# PDF CLASSIFIER
# ────────────────────────────────────────────────────────────

class PDFClassifier:
    """Detect whether PDF is text-based, scanned, or mixed."""

    def __init__(self, logger: logging.Logger, password: Optional[str] = None):
        self.logger = logger
        self.password = password or ""

    def classify(self, pdf_path: str) -> Dict[str, Any]:
        result = {
            "type": "unknown",
            "page_count": 0,
            "text_pages": [],
            "scanned_pages": [],
            "mixed": False,
            "has_tables": False,
            "estimated_tables": 0
        }

        if not PYMUPDF_AVAILABLE:
            self.logger.warning("PyMuPDF not available; assuming text-based PDF.")
            result["type"] = "text"
            return result

        doc = None
        try:
            doc = fitz.open(pdf_path)
            # An encrypted document reports pages but returns no text until it
            # is unlocked, which would classify every page as scanned and send
            # a perfectly readable file to OCR.
            if doc.needs_pass and not doc.authenticate(self.password):
                self.logger.error(
                    f"'{pdf_path}' is password-protected and was not unlocked."
                )
                result["type"] = "encrypted"
                return result

            result["page_count"] = len(doc)

            for page_num, page in enumerate(doc):
                text = page.get_text().strip()
                blocks = page.get_text("blocks")
                images = page.get_images(full=True)

                has_text = len(text) > 50
                has_images = len(images) > 0
                # Heuristic: if images cover >60% of page and little text ──────── scanned
                if has_images and not has_text:
                    result["scanned_pages"].append(page_num)
                elif has_text:
                    result["text_pages"].append(page_num)
                    # Rough table detection via block grid pattern
                    if len(blocks) > 10:
                        result["estimated_tables"] += 1
                else:
                    result["scanned_pages"].append(page_num)

            tp = len(result["text_pages"])
            sp = len(result["scanned_pages"])
            total = result["page_count"]

            if total == 0:
                result["type"] = "empty"
            elif sp == 0:
                result["type"] = "text"
            elif tp == 0:
                result["type"] = "scanned"
            else:
                result["type"] = "mixed"
                result["mixed"] = True

            result["has_tables"] = result["estimated_tables"] > 0
            self.logger.info(
                f"Classified '{pdf_path}': {result['type']}, "
                f"{total} pages ({tp} text, {sp} scanned)"
            )
        except Exception as e:
            self.logger.error(f"Classification failed for '{pdf_path}': {e}")
            result["type"] = "text"  # fallback
        finally:
            # Closed here rather than at the end of the try: an exception part
            # way through the page loop used to leak the open document, which
            # on Windows leaves the file locked for the rest of the run.
            if doc is not None:
                try:
                    doc.close()
                except Exception:
                    pass

        return result


# ────────────────────────────────────────────────────────────
# OCR PROCESSOR
# ────────────────────────────────────────────────────────────

class OCRProcessor:
    """Convert scanned PDF pages to text/structured data via Tesseract."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger

        # Guarded on availability, not just on the setting. `pytesseract` is an
        # optional import, so reading its attributes when the package is absent
        # raised NameError here - before any conversion started, and only for
        # users who had followed the README's advice to set tesseract_path.
        if config.get("tesseract_path"):
            if TESSERACT_AVAILABLE:
                pytesseract.pytesseract.tesseract_cmd = config["tesseract_path"]
            else:
                self.logger.warning(
                    "tesseract_path is set in config but pytesseract is not "
                    "installed, so scanned pages will be skipped. "
                    "Install it with: pip install pytesseract"
                )

    def page_to_image(
        self, pdf_path: str, page_num: int,
        handles: Optional["_PDFHandles"] = None
    ) -> Optional[Any]:
        if not PYMUPDF_AVAILABLE or not PILLOW_AVAILABLE:
            return None
        if handles is not None:
            return self._render_page(handles.fitz(), page_num)
        with _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password")) as own:
            return self._render_page(own.fitz(), page_num)

    def _render_page(self, doc: Optional[Any], page_num: int) -> Optional[Any]:
        if doc is None:
            return None
        try:
            page = doc[page_num]
            dpi = self.config.get("ocr_dpi", 300)
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, alpha=False)
            img_bytes = pix.tobytes("png")
            img = Image.open(io.BytesIO(img_bytes))
            return self._preprocess_image(img)
        except Exception as e:
            self.logger.error(f"Page-to-image failed (page {page_num}): {e}")
            return None

    def _preprocess_image(self, img: Any) -> Any:
        """Enhance image quality for better OCR accuracy."""
        img = img.convert("L")  # grayscale
        enhancer = ImageEnhance.Contrast(img)
        img = enhancer.enhance(2.0)
        img = img.filter(ImageFilter.SHARPEN)
        return img

    def extract_text(self, img: Any) -> str:
        if not TESSERACT_AVAILABLE:
            self.logger.warning("pytesseract not available; skipping OCR.")
            return ""
        try:
            lang = self.config.get("ocr_language", "eng")
            config_str = "--oem 3 --psm 6"
            text = pytesseract.image_to_string(img, lang=lang, config=config_str)
            return text
        except Exception as e:
            self.logger.error(f"OCR text extraction failed: {e}")
            return ""

    def extract_table_data(self, img: Any) -> Optional[pd.DataFrame]:
        """Use Tesseract's TSV output to reconstruct table structure."""
        if not TESSERACT_AVAILABLE:
            return None
        try:
            lang = self.config.get("ocr_language", "eng")
            config_str = "--oem 3 --psm 6"
            threshold = self.config.get("ocr_confidence_threshold", 60)

            data = pytesseract.image_to_data(
                img, lang=lang, config=config_str,
                output_type=pytesseract.Output.DATAFRAME
            )
            data = data[data["conf"] >= threshold].copy()
            data = data[data["text"].notna() & (data["text"].str.strip() != "")]

            if data.empty:
                return None

            # Group by line_num and block_num to reconstruct rows
            rows = []
            for (block, par, line), group in data.groupby(["block_num", "par_num", "line_num"]):
                row_text = " ".join(group["text"].astype(str).tolist())
                rows.append({"block": block, "line": line, "text": row_text})

            if not rows:
                return None

            df = pd.DataFrame(rows)
            # Try to detect tabular structure by splitting on consistent whitespace
            return self._parse_text_table(df["text"].tolist())
        except Exception as e:
            self.logger.error(f"OCR table extraction failed: {e}")
            return None

    def _parse_text_table(self, lines: List[str]) -> Optional[pd.DataFrame]:
        """Parse space-aligned text lines into a DataFrame."""
        if not lines:
            return None
        try:
            # Use pandas read_fwf on the text block
            text_block = "\n".join(lines)
            buf = io.StringIO(text_block)
            df = pd.read_fwf(buf, header=None)
            df = df.dropna(how="all").reset_index(drop=True)
            return df if not df.empty else None
        except Exception:
            # Fallback: treat each line as a single column
            return pd.DataFrame({"text": lines})


# ────────────────────────────────────────────────────────────
# TABLE EXTRACTOR
# ────────────────────────────────────────────────────────────

class TableExtractor:
    """Try multiple extraction methods and return best results."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.method = config.get("table_extraction_method", "auto")

    def extract(self, pdf_path: str, page_nums: List[int]) -> List[Dict]:
        """Returns list of {page, table_index, dataframe, source}."""
        results = []

        # One open document for the whole page loop instead of one per page.
        with _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password")) as handles:
            for page_num in page_nums:
                tables = self._extract_page(pdf_path, page_num, handles)
                for i, df in enumerate(tables):
                    if df is not None and not df.empty:
                        results.append({
                            "page": page_num + 1,
                            "table_index": i + 1,
                            "dataframe": df,
                            "source": "table_extractor"
                        })

        return results

    def _extract_page(
        self, pdf_path: str, page_num: int,
        handles: Optional["_PDFHandles"] = None
    ) -> List[pd.DataFrame]:
        method = self.method

        if method == "auto":
            # Try pdfplumber first (best for native PDFs)
            tables = self._try_pdfplumber(pdf_path, page_num, handles)
            if tables:
                self.logger.debug(f"Page {page_num+1}: pdfplumber found {len(tables)} tables")
                return tables

            # Try camelot (great for complex tables)
            if CAMELOT_AVAILABLE:
                tables = self._try_camelot(pdf_path, page_num)
                if tables:
                    self.logger.debug(f"Page {page_num+1}: camelot found {len(tables)} tables")
                    return tables

            # Try tabula (Java-based, robust)
            if TABULA_AVAILABLE:
                tables = self._try_tabula(pdf_path, page_num)
                if tables:
                    self.logger.debug(f"Page {page_num+1}: tabula found {len(tables)} tables")
                    return tables

            return []

        elif method == "pdfplumber":
            return self._try_pdfplumber(pdf_path, page_num, handles)
        elif method == "camelot" and CAMELOT_AVAILABLE:
            return self._try_camelot(pdf_path, page_num)
        elif method == "tabula" and TABULA_AVAILABLE:
            return self._try_tabula(pdf_path, page_num)
        else:
            return self._try_pdfplumber(pdf_path, page_num, handles)

    def _try_pdfplumber(
        self, pdf_path: str, page_num: int,
        handles: Optional["_PDFHandles"] = None
    ) -> List[pd.DataFrame]:
        if not PDFPLUMBER_AVAILABLE:
            return []
        if handles is not None:
            return self._pdfplumber_page_tables(handles.plumber(), page_num)
        with _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password")) as own:
            return self._pdfplumber_page_tables(own.plumber(), page_num)

    def _pdfplumber_page_tables(
        self, pdf: Optional[Any], page_num: int
    ) -> List[pd.DataFrame]:
        if pdf is None:
            return []
        try:
            if page_num >= len(pdf.pages):
                return []
            page = pdf.pages[page_num]
            try:
                tables = page.extract_tables()
            finally:
                _flush_page_cache(page)
            result = []
            for t in tables:
                if t:
                    df = pd.DataFrame(t)
                    df = self._clean_dataframe(df, promote_header=True)
                    if not df.empty:
                        result.append(df)
            return result
        except Exception as e:
            self.logger.debug(f"pdfplumber failed on page {page_num+1}: {e}")
            return []

    def _try_camelot(self, pdf_path: str, page_num: int) -> List[pd.DataFrame]:
        try:
            page_str = str(page_num + 1)
            # Try lattice (grid lines) first, then stream (whitespace)
            for flavor in ("lattice", "stream"):
                try:
                    tables = camelot.read_pdf(pdf_path, pages=page_str, flavor=flavor)
                    if tables.n > 0:
                        result = []
                        for t in tables:
                            df = self._clean_dataframe(t.df, promote_header=True)
                            if not df.empty:
                                result.append(df)
                        return result
                except Exception:
                    continue
            return []
        except Exception as e:
            self.logger.debug(f"camelot failed on page {page_num+1}: {e}")
            return []

    def _try_tabula(self, pdf_path: str, page_num: int) -> List[pd.DataFrame]:
        try:
            dfs = tabula.read_pdf(
                pdf_path, pages=page_num + 1,
                multiple_tables=True, silent=True
            )
            result = []
            for df in (dfs or []):
                df = self._clean_dataframe(df, promote_header=False)
                if not df.empty:
                    result.append(df)
            return result
        except Exception as e:
            self.logger.debug(f"tabula failed on page {page_num+1}: {e}")
            return []

    def _clean_dataframe(
        self, df: pd.DataFrame,
        promote_header: bool = True
    ) -> pd.DataFrame:
        if df is None or df.empty:
            return pd.DataFrame()

        df = df.copy()

        # Replace None/empty strings with NaN
        df = df.replace(r"^\s*$", np.nan, regex=True)
        df = df.replace({None: np.nan})

        if self.config.get("remove_empty_rows", True):
            df = df.dropna(how="all")

        if self.config.get("remove_empty_cols", True):
            df = df.dropna(axis=1, how="all")

        if df.empty:
            return df

        # Promote first row to header if it looks like a header
        if promote_header and not df.empty:
            first_row = df.iloc[0].astype(str).tolist()
            is_header = all(not s.replace(".", "").replace("-", "").isdigit()
                            for s in first_row if s and s != "nan")
            if is_header:
                df.columns = first_row
                df = df.iloc[1:].reset_index(drop=True)

        df = self._merge_split_report_rows(df)
        df = self._normalize_item_columns(df)
        df = df.reset_index(drop=True)
        return df

    def _merge_split_report_rows(self, df: pd.DataFrame) -> pd.DataFrame:
        expected_cols = {
            "s. no.", "customer name", "item", "product type",
            "certificate no", "order no", "amount",
            "manufacturer", "est shipping date"
        }
        normalized_cols = [str(col).strip().lower() for col in df.columns]
        if not expected_cols.issubset(set(normalized_cols)):
            return df

        rows: List[Dict[str, Any]] = []
        last_row: Optional[Dict[str, Any]] = None

        for _, raw_row in df.iterrows():
            row = {str(col): raw_row[col] for col in df.columns}
            non_empty = {
                str(col): raw_row[col]
                for col in df.columns
                if not pd.isna(raw_row[col]) and str(raw_row[col]).strip() != ""
            }
            if not non_empty:
                continue

            sno_val = row.get("S. No.")
            has_sno = not pd.isna(sno_val) and str(sno_val).strip() != ""

            if has_sno or last_row is None:
                rows.append(row)
                last_row = rows[-1]
                continue

            if len(non_empty) == 1 and "Certificate No" in non_empty:
                merged = " ".join([
                    str(last_row.get("Certificate No", "")).strip(),
                    str(non_empty["Certificate No"]).strip(),
                ]).strip()
                last_row["Certificate No"] = re.sub(r"\s+", " ", merged)
                continue

            if len(non_empty) == 1 and "Item" in non_empty:
                merged = " ".join([
                    str(last_row.get("Item", "")).strip(),
                    str(non_empty["Item"]).strip(),
                ]).strip()
                last_row["Item"] = re.sub(r"\s+", " ", merged)
                continue

            for key, value in non_empty.items():
                existing = last_row.get(key)
                if pd.isna(existing) or str(existing).strip() == "":
                    last_row[key] = value
                elif key in {"Certificate No", "Item"}:
                    last_row[key] = re.sub(
                        r"\s+", " ", f"{str(existing).strip()} {str(value).strip()}"
                    ).strip()

        if not rows:
            return df

        return pd.DataFrame(rows, columns=df.columns)

    def _normalize_item_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        for column in df.columns:
            if str(column).strip().lower() != "item":
                continue
            df[column] = df[column].apply(self._normalize_item_value)
        return df

    def _normalize_item_value(self, value: Any) -> Any:
        if pd.isna(value):
            return value

        text = str(value)
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        compact = re.sub(r"\s+", " ", " ".join(lines)).strip()
        return compact


# ────────────────────────────────────────────────────────────
# REPORT COLUMN PARSER (geometry-based)
# ────────────────────────────────────────────────────────────

# The report this tool was built for. Kept in code as the always-available
# fallback, so adding layouts in config can never stop it being read.
DEFAULT_REPORT_LAYOUT: Dict[str, Any] = {
    "name": "supplier_order_report",
    "columns": [
        {"name": "S. No.", "tokens": ["S.", "No."]},
        {"name": "Customer Name", "tokens": ["Customer", "Name"]},
        {"name": "Item", "tokens": ["Item"]},
        {"name": "Product Type", "tokens": ["Product", "Type"]},
        {"name": "Certificate No", "tokens": ["Certificate", "No"]},
        {"name": "Order No", "tokens": ["Order", "No"]},
        {"name": "Amount", "tokens": ["Amount"]},
        {"name": "Manufacturer", "tokens": ["Manufacturer"]},
        {"name": "Est Shipping Date", "tokens": ["Est", "Shipping", "Date"]},
    ],
    "row_key": "S. No.",
    "row_key_pattern": r"^\d+$",
    # Single-line columns a wrapped descriptive cell never fills. A wrap only
    # ever continues a descriptive column; it does not restate an Amount or a
    # shipping date. So a "continuation" line carrying either is not a wrap - it
    # is a footer, or a data row whose serial did not extract. Merging it would
    # silently corrupt or absorb a record.
    "structural_columns": ["Amount", "Est Shipping Date"],
    # Certificate No and Product Type are deliberately absent: settings and
    # mountings legitimately have no certificate.
    "required": ["Order No"],
    "patterns": {
        "Amount": r"^\$?\d[\d,]*\.\d{2}$",
        "Est Shipping Date": r"^\d{1,2}\s+[A-Za-z]{3}\s+\d{4}$",
    },
    # Identifiers stay text so Excel cannot renumber or re-date them; the
    # serial is a counter, and the amount is money.
    "formats": {
        "S. No.": "integer",
        "Order No": "text",
        "Certificate No": "text",
        "Amount": "money",
    },
}


# Friendly names for the Excel number formats a report actually needs. Anything
# not listed is passed through as a literal Excel format string, so an unusual
# case is possible without this table having to anticipate it.
NUMBER_FORMAT_ALIASES: Dict[str, str] = {
    "money": "$#,##0.00",
    "number": "#,##0.00",
    "integer": "#,##0",
    "percent": "0.0%",
    "text": "@",
}


class ReportLayout:
    """One report shape: its columns, and what makes a row of it valid.

    Both halves belong together. Describing a layout's columns without its
    validation rules would let a new report parse geometrically and then have
    every row rejected by rules written for a different document - the tool's
    loudest failure mode firing on a layout that is actually fine.
    """

    def __init__(self, spec: Dict[str, Any]):
        self.name = str(spec.get("name") or "unnamed")

        self.columns: List[Tuple[str, Tuple[str, ...]]] = []
        for entry in spec.get("columns") or []:
            try:
                name = str(entry["name"])
                tokens = tuple(str(t) for t in entry["tokens"])
            except (TypeError, KeyError) as e:
                raise ValueError(f"each column needs 'name' and 'tokens' ({e})")
            if not name or not tokens:
                raise ValueError("column 'name' and 'tokens' cannot be empty")
            self.columns.append((name, tokens))

        if not self.columns:
            raise ValueError("layout defines no columns")

        names = [n for n, _ in self.columns]
        duplicates = {n for n in names if names.count(n) > 1}
        if duplicates:
            raise ValueError(f"duplicate column names: {sorted(duplicates)}")

        def _known(values, field):
            unknown = [v for v in values if v not in names]
            if unknown:
                raise ValueError(f"{field} names unknown columns: {unknown}")
            return tuple(values)

        self.row_key = str(spec.get("row_key") or names[0])
        if self.row_key not in names:
            raise ValueError(f"row_key '{self.row_key}' is not one of the columns")

        try:
            self.row_key_pattern = re.compile(
                str(spec.get("row_key_pattern") or r"^\d+$")
            )
            self.patterns = {
                col: re.compile(str(pat))
                for col, pat in (spec.get("patterns") or {}).items()
            }
        except re.error as e:
            raise ValueError(f"invalid regular expression: {e}")

        _known(list(self.patterns), "patterns")
        self.structural_columns = _known(
            list(spec.get("structural_columns") or []), "structural_columns")
        self.required = _known(list(spec.get("required") or []), "required")

        # How each column is written to Excel. A price read as "4.50" is the
        # number 4.5, and without a declared format Excel shows it as "4.5" -
        # right value, wrong for money. Declaring it here drives both the cell
        # format and how the string is converted, so the two cannot disagree.
        raw_formats = spec.get("formats") or {}
        _known(list(raw_formats), "formats")
        self.formats: Dict[str, str] = {
            col: NUMBER_FORMAT_ALIASES.get(str(fmt), str(fmt))
            for col, fmt in raw_formats.items()
        }

    def column_names(self) -> List[str]:
        return [n for n, _ in self.columns]

    def is_row_start(self, joined: Dict[str, str]) -> bool:
        """True when this line begins a new record rather than continuing one."""
        return bool(self.row_key_pattern.match(joined.get(self.row_key, "").strip()))

    def row_is_complete(self, row: Dict[str, str]) -> bool:
        """A row must carry the fields that make it a usable record."""
        if not self.is_row_start(row):
            return False
        for col in self.required:
            if not str(row.get(col, "")).strip():
                return False
        for col, pattern in self.patterns.items():
            if not pattern.match(str(row.get(col, "")).strip()):
                return False
        return True


def load_report_layouts(
    config: Optional[dict], logger: logging.Logger
) -> List[ReportLayout]:
    """Build the layouts to try, from config plus the built-in one.

    Configured layouts are tried first so they can take precedence, and the
    built-in is always appended last - adding a layout can therefore never stop
    the report this tool was written for from being read. A malformed entry is
    reported and skipped rather than taking the whole run down with it.
    """
    layouts: List[ReportLayout] = []
    for i, spec in enumerate((config or {}).get("report_layouts") or []):
        try:
            layouts.append(ReportLayout(spec))
        except Exception as e:
            logger.warning(
                f"config report_layouts[{i}] ignored - {e}. "
                f"The other layouts are unaffected."
            )

    layouts.append(ReportLayout(DEFAULT_REPORT_LAYOUT))
    if len(layouts) > 1:
        logger.info(
            "Report layouts in use: "
            + ", ".join(lay.name for lay in layouts)
        )
    return layouts


class ReportColumnParser:
    """Read a tabular report from the PDF's own column geometry.

    Matching patterns against the flattened text guesses at structure and fails
    whenever a cell is blank, wraps, or holds an unexpected value. Word
    x-positions state the structure outright: a value belongs to whichever
    column it is physically printed under. That handles blank certificates,
    unexpected product types and wrapped cells with no special-casing.

    Column boundaries are calibrated per page from the widest empty vertical
    strip between adjacent header labels, so they follow the document rather
    than hard-coded coordinates.
    """

    LINE_TOLERANCE = 3.0     # points; words within this vertical span are one line

    def __init__(
        self, logger: logging.Logger,
        layouts: Optional[List[ReportLayout]] = None
    ):
        self.logger = logger
        self.layouts = layouts or [ReportLayout(DEFAULT_REPORT_LAYOUT)]
        # Lines that could be attached to no row during the last parse_page():
        # unrecognised footers, or data rows whose serial did not read. The
        # caller surfaces these so nothing is dropped silently.
        self.unparsed_lines: List[str] = []
        # Which layout matched the last page parsed, so the caller can label
        # the result and validate rows by that layout's own rules.
        self.layout: Optional[ReportLayout] = None

    # ── line grouping ────────────────────────────────────────
    def _visual_lines(self, words: List[Dict]) -> List[List[Dict]]:
        lines: List[List[Dict]] = []
        for w in sorted(words, key=lambda w: (w["top"], w["x0"])):
            if lines and abs(w["top"] - lines[-1][0]["top"]) <= self.LINE_TOLERANCE:
                lines[-1].append(w)
            else:
                lines.append([w])
        return [sorted(ln, key=lambda w: w["x0"]) for ln in lines]

    # ── header location ──────────────────────────────────────
    def _match_header(
        self, lines: List[List[Dict]]
    ) -> Optional[Tuple[int, ReportLayout, List[float]]]:
        """Find the first line that is a header, and say which layout it is.

        The header test *is* the anchor test: a line qualifies only if it
        carries every column label, as whole words, consecutively and in order.
        That is strict enough to identify the layout outright, so no layout
        needs a separate hand-written "does this look like my header" rule.
        """
        for i, ln in enumerate(lines):
            for layout in self.layouts:
                anchors = self._header_anchors(ln, layout)
                if anchors is not None:
                    return i, layout, anchors
        return None

    def _header_anchors(
        self, header: List[Dict], layout: ReportLayout
    ) -> Optional[List[float]]:
        """Centre x of each column label, in column order."""
        anchors: List[float] = []
        idx = 0
        for _name, tokens in layout.columns:
            start = idx
            for tok in tokens:
                if idx >= len(header) or header[idx]["text"] != tok:
                    return None
                idx += 1
            anchors.append((header[start]["x0"] + header[idx - 1]["x1"]) / 2.0)
        return anchors

    # ── boundary calibration ─────────────────────────────────
    def _boundaries(self, anchors: List[float], body: List[List[Dict]]) -> List[float]:
        """Split adjacent columns at the widest gap in the printed data.

        Headers are centred over their columns while data is left-aligned, so a
        naive midpoint between header labels lands inside real values. The
        widest empty strip between two anchors is where the column actually ends.
        """
        spans = [(w["x0"], w["x1"]) for ln in body for w in ln]
        bounds: List[float] = []
        for left, right in zip(anchors, anchors[1:]):
            best_start, best_len = None, -1.0
            x, step = left, 0.5
            run_start = None
            while x <= right:
                covered = any(a - 0.5 <= x <= b + 0.5 for a, b in spans)
                if not covered:
                    if run_start is None:
                        run_start = x
                else:
                    if run_start is not None and (x - run_start) > best_len:
                        best_start, best_len = run_start, x - run_start
                    run_start = None
                x += step
            if run_start is not None and (right - run_start) > best_len:
                best_start, best_len = run_start, right - run_start
            bounds.append((best_start + best_len / 2.0) if best_start is not None
                          else (left + right) / 2.0)
        return bounds

    def _is_non_row(self, line: List[Dict], layout: ReportLayout) -> bool:
        """True for a repeated header or a recognised footer line.

        Used both to skip these lines when reading rows and to keep them out of
        the boundary calibration, so the two can never disagree.

        A repeated header is detected by running the anchor test again rather
        than by matching a hard-coded opening string, so this works for any
        layout without each one having to describe its header twice.
        """
        if self._header_anchors(line, layout) is not None:
            return True
        text = " ".join(w["text"] for w in line).strip()
        return bool(TextExtractor.FOOTER_MARKER.match(text))

    def _column_of(self, x: float, bounds: List[float]) -> int:
        col = 0
        for b in bounds:
            if x >= b:
                col += 1
            else:
                break
        return col

    # ── main entry ───────────────────────────────────────────
    def parse_page(self, page: Any) -> Optional[List[Dict[str, str]]]:
        # Cleared here rather than beside the row loop below: every early return
        # in between would otherwise leave the previous page's lines in place,
        # and the caller reads this list after the call.
        self.unparsed_lines = []
        self.layout = None

        try:
            words = page.extract_words()
        except Exception as e:
            self.logger.debug(f"extract_words failed: {e}")
            return None
        if not words:
            return None

        lines = self._visual_lines(words)
        match = self._match_header(lines)
        if match is None:
            return None

        hdr_i, layout, anchors = match
        self.layout = layout

        body = lines[hdr_i + 1:]
        if not body:
            return []

        # Calibrate from data rows only. Footers and repeated headers do not
        # sit in the table's columns, and their words can land squarely in the
        # empty strip that separates two columns - "Total: $19,824.50" printed
        # under the first column fills the gap between S. No. and Customer
        # Name, moving the boundary right and swallowing the first name into
        # the serial column, which rejects every row on the page. The row loop
        # below already skips these lines; the geometry has to skip them too.
        calibration = [ln for ln in body if not self._is_non_row(ln, layout)]
        bounds = self._boundaries(anchors, calibration or body)
        names = layout.column_names()

        rows: List[Dict[str, str]] = []
        for ln in body:
            text = " ".join(w["text"] for w in ln).strip()
            # Repeated header on a later page, or a summary/footer line.
            if self._is_non_row(ln, layout):
                continue

            cells = {n: [] for n in names}
            for w in ln:
                cells[names[self._column_of(w["x0"], bounds)]].append(w["text"])
            joined = {n: " ".join(v).strip() for n, v in cells.items()}

            if layout.is_row_start(joined):
                rows.append(joined)
                continue

            # Not a new row. It is only a wrapped continuation of the row above
            # if it carries no single-line structural value (see the layout's
            # structural_columns). Otherwise it is a footer we did not
            # recognise, or a data row whose key did not read — record it rather
            # than merge it, so a real record can never be silently glued onto
            # its predecessor.
            structural = any(joined[c] for c in layout.structural_columns)
            if rows and not structural:
                for n in names:
                    if joined[n]:
                        rows[-1][n] = (rows[-1][n] + " " + joined[n]).strip()
            elif text:
                self.unparsed_lines.append(text)

        return rows


# ────────────────────────────────────────────────────────────
# TEXT EXTRACTOR
# ────────────────────────────────────────────────────────────

class TextExtractor:
    """Extract raw text content page by page."""

    # The flat-text fallback below is specific to the built-in layout: it is a
    # last resort for when the page geometry cannot be read at all, and a
    # regex per layout would be a far weaker thing than the column parser that
    # already handles them. Configured layouts are read by geometry; if that
    # fails for one, its pages fall through to generic text extraction rather
    # than being matched by these patterns.
    REPORT_HEADER = (
        "S. No. Customer Name Item Product Type Certificate No "
        "Order No Amount Manufacturer Est Shipping Date"
    )

    # Summary/footer content that trails the table body. Checked before the
    # new-row test, because lines like "95 Items - $94222.52" start with digits
    # and would otherwise be mistaken for a data row.
    FOOTER_MARKER = re.compile(
        r"^\s*(?:"
        r"(?:grand\s+|sub[\s-]?)?total\b"
        r"|amount\s+due\b|balance\s+due\b|invoice\s+total\b"
        r"|page\s+\d+\b"
        r"|thank\s+you\b"
        r"|payment\s+terms\b|terms\b"
        r"|generated\s+(?:on|by)\b|printed\s+(?:on|by)\b"
        r"|\d[\d,]*\s+items?\b"
        r"|signature\b|authori[sz]ed\b"
        r"|continued\b"
        r")",
        re.IGNORECASE,
    )

    # A wrapped row continues onto the next line as a short run of these
    # fragments — clarity/colour grades, shapes, carat weights, lab names and
    # certificate ids. Anything else is footer or boilerplate and must not be
    # appended to the previous row.
    CONTINUATION_TOKEN = re.compile(
        r"^(?:"
        r"lg[_-]?\d+"
        r"|igi|gia|ags|hrd|egl"
        r"|\d{6,}"
        r"|fl|if|vvs1|vvs2|vs1|vs2|si1|si2|si3|i1|i2|i3"
        # Cut grades. These head the wrapped line on many rows, e.g.
        # "Excellent LG_100000002" — omitting them stranded the certificate id.
        r"|excellent|ideal|astor|premium|super|very|good|fair|poor"
        r"|[d-z]"
        r"|round|oval|cushion|emerald|princess|pear|marquise"
        r"|asscher|radiant|heart|trillion|baguette"
        r"|\d+(?:\.\d+)?"
        r")$",
        re.IGNORECASE,
    )

    # Wrapped fragments are short; prose and footers are not.
    MAX_CONTINUATION_TOKENS = 4

    def __init__(self, logger: logging.Logger, config: Optional[dict] = None):
        self.logger = logger
        self.config = config or {}
        # Report lines that matched no known row shape during the last extract().
        self.dropped_rows: List[str] = []
        self.column_parser = ReportColumnParser(
            logger, load_report_layouts(config, logger))
        # Layout names that actually read something, surfaced in the workbook's
        # metadata so the output says which shape it was read as.
        self.matched_layouts: List[str] = []
        # Column formats declared by the layout that matched, carried through
        # to the Excel writer so a configured money column is written as money.
        self.column_formats: Dict[str, str] = {}

    def extract(self, pdf_path: str, page_nums: List[int]) -> List[Dict]:
        results = []
        self.dropped_rows = []
        # One open document for the whole page loop instead of one per page.
        with _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password")) as handles:
            pdf = handles.plumber()
            for page_num in page_nums:
                text = self._extract_page_text(pdf_path, page_num, handles)

                # Prefer the report's own column geometry; fall back to parsing
                # the flattened text when the layout cannot be read.
                df = self._report_by_columns(pdf, page_num)
                if df is not None and not df.empty:
                    results.append({
                        "page": page_num + 1,
                        "text": text,
                        "dataframe": df
                    })
                    continue

                # An empty (rather than None) frame means the layout was read and
                # every row it held has already been reported as excluded, so the
                # report text parser must not run over this page a second time.
                layout_was_read = df is not None

                if text:
                    results.append({
                        "page": page_num + 1,
                        "text": text,
                        "dataframe": self._text_to_dataframe(
                            text, parse_report=not layout_was_read
                        )
                    })

        if self.dropped_rows:
            self.logger.warning(
                f"{len(self.dropped_rows)} report row(s) did not match the expected "
                f"format and were excluded from the output. See warnings above."
            )
        return results

    def _report_by_columns(self, pdf: Optional[Any], page_num: int) -> Optional[pd.DataFrame]:
        """Parse a report page positionally, validating each row before keeping it.

        The return value carries two outcomes the caller must treat differently:

        * ``None`` - the page's column layout could not be read at all, so the
          caller falls back to parsing the flattened text.
        * an **empty** DataFrame - the layout *was* read, but no row survived
          validation. Every exclusion has already been reported by this method,
          so the caller must not re-run the report text parser over the same
          page. Doing so counted each excluded row twice, inflating the figure
          shown in the metadata sheet and the batch summary.
        """
        if pdf is None:
            return None
        try:
            if page_num >= len(pdf.pages):
                return None
            page = pdf.pages[page_num]
            try:
                rows = self.column_parser.parse_page(page)
            finally:
                _flush_page_cache(page)
        except Exception as e:
            self.logger.debug(f"Column parse failed on page {page_num+1}: {e}")
            return None

        # None means no readable header; [] means a header with an empty body,
        # which is still a layout this parser understood.
        if rows is None:
            return None

        # Validated by the rules of the layout that actually matched, not by a
        # single fixed rule. Judging one report by another's required fields
        # would reject every row of a perfectly good page and report the lot as
        # excluded - the loudest failure this tool has, fired at nothing.
        layout = self.column_parser.layout
        if layout is None:
            return None

        kept = []
        for row in rows:
            if layout.row_is_complete(row):
                if "Item" in row:
                    row["Item"] = self._normalize_item_whitespace(row["Item"])
                kept.append(row)
            else:
                self._note_dropped_row(
                    " ".join(f"{k}={v}" for k, v in row.items() if v)
                )

        # Surfaced before the empty-result check below, not after it. A line the
        # parser could attach to no row (an unrecognised footer, or a data row
        # whose serial did not read) must be reported even on a page that yields
        # nothing else - otherwise it vanishes without a trace, which is the one
        # outcome this product does not accept.
        for line in self.column_parser.unparsed_lines:
            self._note_dropped_row(line)

        columns = layout.column_names()
        if not kept:
            return pd.DataFrame(columns=columns)

        if layout.name not in self.matched_layouts:
            self.matched_layouts.append(layout.name)
        self.column_formats = dict(layout.formats)

        self.logger.info(
            f"Page {page_num+1}: read {len(kept)} row(s) from the "
            f"'{layout.name}' column layout."
        )
        return pd.DataFrame(kept, columns=columns)

    def _extract_page_text(
        self, pdf_path: str, page_num: int,
        handles: Optional["_PDFHandles"] = None
    ) -> str:
        own: Optional[_PDFHandles] = None
        if handles is None:
            own = _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password"))
            handles = own
        try:
            # Prefer pdfplumber for layout-aware extraction
            pdf = handles.plumber()
            if pdf is not None:
                try:
                    if page_num < len(pdf.pages):
                        page = pdf.pages[page_num]
                        try:
                            return page.extract_text() or ""
                        finally:
                            _flush_page_cache(page)
                except Exception as e:
                    self.logger.debug(f"pdfplumber text fail page {page_num+1}: {e}")

            doc = handles.fitz()
            if doc is not None:
                try:
                    return doc[page_num].get_text()
                except Exception as e:
                    self.logger.debug(f"PyMuPDF text fail page {page_num+1}: {e}")

            return ""
        finally:
            if own is not None:
                own.close()

    def _text_to_dataframe(
        self, text: str, parse_report: bool = True
    ) -> pd.DataFrame:
        """Convert raw text to DataFrame, trying to detect structure.

        `parse_report=False` skips the order-report pass for a page the column
        parser already read and already reported on, so its excluded rows are
        not counted a second time here.
        """
        lines = [l.strip() for l in text.split("\n") if l.strip()]
        if not lines:
            return pd.DataFrame()

        if parse_report:
            report_df = self._parse_order_report(lines)
            if not report_df.empty:
                return report_df

        # Detect delimiter-separated structure
        for delimiter in ["\t", "|", ",", "  "]:
            sample = [l.split(delimiter) for l in lines[:5]]
            col_counts = [len(r) for r in sample]
            if min(col_counts) > 1 and max(col_counts) - min(col_counts) <= 1:
                rows = [l.split(delimiter) for l in lines]
                max_cols = max(len(r) for r in rows)
                rows = [r + [""] * (max_cols - len(r)) for r in rows]
                df = pd.DataFrame(rows[1:], columns=rows[0]) if rows else pd.DataFrame()
                if not df.empty:
                    return df

        # Fallback: one column
        return pd.DataFrame({"Content": lines})

    def _parse_order_report(self, lines: List[str]) -> pd.DataFrame:
        header_idx = next(
            (i for i, line in enumerate(lines) if self.REPORT_HEADER in line),
            None
        )
        if header_idx is None:
            return pd.DataFrame()

        grouped_rows: List[Dict[str, Any]] = []
        current: Optional[Dict[str, Any]] = None
        ignored: List[str] = []

        for line in lines[header_idx + 1:]:
            if self.FOOTER_MARKER.match(line):
                # Totals, page numbers, terms — never row data.
                ignored.append(line)
            elif re.match(r"^\d+\s+", line):
                if current:
                    grouped_rows.append(current)
                current = {"main": line, "extra": []}
            elif current is not None and self._is_continuation_line(line):
                current["extra"].append(line)
            else:
                # Unrecognised trailing content. The previous behaviour appended
                # every such line to the preceding row's Item, which silently
                # corrupted the last row of every report with footer text.
                ignored.append(line)

        if current:
            grouped_rows.append(current)

        if ignored:
            sample = "; ".join(ignored[:5])
            self.logger.debug(
                f"Skipped {len(ignored)} non-row line(s) below the report header: {sample}"
            )

        parsed_rows = []
        for row in grouped_rows:
            parsed = self._parse_order_row(row["main"], row["extra"])
            if parsed:
                parsed_rows.append(parsed)
            else:
                self._note_dropped_row(row["main"])

        if grouped_rows and not parsed_rows:
            self.logger.warning(
                f"Report header found but none of the {len(grouped_rows)} data row(s) "
                f"matched the expected format. The report layout may have changed; "
                f"falling back to generic text parsing."
            )

        return pd.DataFrame(parsed_rows)

    def _is_continuation_line(self, line: str) -> bool:
        """True if `line` is the wrapped tail of the preceding report row.

        Only short runs of recognised diamond/certificate fragments qualify.
        Everything else is treated as footer content and skipped, so that a
        line like "Total: $94,222.52" can never be glued onto a row's Item.
        """
        tokens = line.split()
        if not tokens or len(tokens) > self.MAX_CONTINUATION_TOKENS:
            return False
        return all(self.CONTINUATION_TOKEN.match(token) for token in tokens)

    def _note_dropped_row(self, line: str) -> None:
        """Record a report row that could not be parsed.

        These rows are excluded from the output entirely, so they must never be
        dropped silently — a single unrecognised product type or date format
        would otherwise erase real orders from the Excel with no trace.
        """
        self.dropped_rows.append(line)
        preview = line if len(line) <= 160 else line[:157] + "..."
        self.logger.warning(f"Unparsed report row (excluded from output): {preview}")

    def _parse_order_row(
        self, main_line: str, extra_lines: List[str]
    ) -> Optional[Dict[str, Any]]:
        pattern = re.compile(
            r"^(?P<sno>\d+)\s+"
            r"(?P<customer>.+?)\s+"
            r"(?P<item>\d+(?:\.\d+)?\s+.+?)\s+"
            r"(?P<product_type>Diamond)\s+"
            r"(?P<certificate_no>.+?)\s+"
            r"(?P<order_no>[A-Z0-9]{6,})\s+"
            r"(?P<amount>\$\d[\d,]*\.\d{2})\s+"
            r"(?P<manufacturer>.+?)\s+"
            r"(?P<est_shipping_date>\d{2}\s+[A-Za-z]{3}\s+\d{4})$"
        )
        match = pattern.match(main_line)
        if not match:
            return None

        record = match.groupdict()
        record["item"] = self._normalize_item_whitespace(record["item"])

        extra_text = " ".join(extra_lines)
        item_tail_parts = []
        for line in extra_lines:
            cleaned = re.sub(r"\bLG[_-]?\d+\b", "", line, flags=re.IGNORECASE)
            cleaned = re.sub(r"\s+", " ", cleaned).strip()
            if cleaned:
                item_tail_parts.append(cleaned)
        if item_tail_parts:
            record["item"] = re.sub(
                r"\s+", " ", f"{record['item']} {' '.join(item_tail_parts)}"
            ).strip()

        lg_match = re.search(r"\bLG[_-]?\d+\b", extra_text)
        if lg_match:
            record["certificate_no"] = (
                f"{record['certificate_no']} {lg_match.group(0)}"
            ).strip()

        record["s_no"] = record.pop("sno")
        return {
            "S. No.": record["s_no"],
            "Customer Name": record["customer"],
            "Item": record["item"],
            "Product Type": record["product_type"],
            "Certificate No": record["certificate_no"],
            "Order No": record["order_no"],
            "Amount": record["amount"],
            "Manufacturer": record["manufacturer"],
            "Est Shipping Date": record["est_shipping_date"],
        }

    def _normalize_item_whitespace(self, text: str) -> str:
        """Collapse whitespace in an item value.

        The column parser joins a wrapped cell with spaces and newlines, so the
        reassembled item can carry runs of whitespace that are tidied here.
        """
        return re.sub(r"\s+", " ", str(text)).strip()


# ────────────────────────────────────────────────────────────
# EXCEL BUILDER
# ────────────────────────────────────────────────────────────

class ExcelBuilder:
    """Build formatted .xlsx from extracted data."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.wb = Workbook()
        # Remove default empty sheet
        self.wb.remove(self.wb.active)
        self._sheet_names: List[str] = []

    def _unique_sheet_name(self, name: str) -> str:
        max_len = self.config.get("max_sheet_name_length", 31)
        # Sanitize invalid Excel sheet name chars
        name = re.sub(r'[\\/:*?\[\]]', '_', name)
        name = name[:max_len]
        base = name
        counter = 1
        while name in self._sheet_names:
            suffix = f"_{counter}"
            name = base[:max_len - len(suffix)] + suffix
            counter += 1
        self._sheet_names.append(name)
        return name

    def add_dataframe(
        self, df: pd.DataFrame, sheet_name: str,
        title: str = "", formats: Optional[Dict[str, str]] = None
    ) -> None:
        """Write a sheet. `formats` maps column name -> Excel number format.

        A declared format wins over the name-based defaults below, which only
        know the built-in report's column names and cannot know what a
        configured layout calls its money column.
        """
        formats = formats or {}
        if df is None or df.empty:
            return

        ws = self.wb.create_sheet(self._unique_sheet_name(sheet_name))
        row_offset = 1

        if title:
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(bold=True, size=12)
            row_offset = 3

        header_fill = PatternFill(
            "solid", fgColor=self.config.get("header_color", "1F4E79")
        )
        alt_fill = PatternFill(
            "solid", fgColor=self.config.get("alternating_row_color", "D6E4F0")
        )
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = Side(border_style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        cols = list(df.columns)

        # Write header
        for col_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=row_offset, column=col_idx, value=str(col_name))
            if self.config.get("apply_table_style", True):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
                cell.border = border

        if self.config.get("freeze_header_row", True):
            ws.freeze_panes = ws.cell(row=row_offset + 1, column=1)

        # Write data rows
        for r_idx, row_data in enumerate(df.itertuples(index=False), 1):
            for c_idx, value in enumerate(row_data, 1):
                col_name = str(cols[c_idx - 1])
                declared = formats.get(col_name)
                if declared:
                    cell = ws.cell(
                        row=row_offset + r_idx, column=c_idx,
                        value=self._coerce_for_format(value, declared))
                    cell.number_format = declared
                else:
                    cell = ws.cell(row=row_offset + r_idx, column=c_idx,
                                   value=self._coerce_value(value, col_name))
                    self._apply_cell_format(cell, col_name)
                if self.config.get("apply_table_style", True):
                    if r_idx % 2 == 0:
                        cell.fill = alt_fill
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=False)

        # Auto-adjust column widths
        self._auto_size_columns(ws, df, row_offset)

    def _coerce_for_format(self, value: Any, number_format: str) -> Any:
        """Convert a value to suit the format its layout declared for it.

        A money or number format needs an actual number in the cell, or Excel
        shows the raw string and the format does nothing - so "$4,722.30" has
        its currency symbol and separators stripped here. A text format keeps
        the value verbatim, which is what protects identifiers.
        """
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None

        text = str(value).strip()
        if not text:
            return None
        if "@" in number_format:
            return text

        # Strip anything that is decoration rather than magnitude: currency
        # symbols, thousands separators, stray spaces.
        cleaned = re.sub(r"[^\d.\-]", "", text)
        if not cleaned or cleaned in {"-", ".", "-."}:
            return text          # not a number after all; keep what was read
        try:
            number = float(cleaned)
        except ValueError:
            return text

        # A format with no decimal places wants an integer in the cell.
        if "." not in number_format and number == int(number):
            return int(number)
        return number

    def _coerce_value(self, value: Any, col_name: str = "") -> Any:
        if pd.isna(value) if isinstance(value, float) else value is None:
            return None

        normalized_col = col_name.strip().lower()

        # A serial number is a counter, so it stays numeric: Excel then sorts it
        # as a number and stops flagging "number stored as text" on the column.
        # Every other name below is an identifier or free text that Excel must
        # not reinterpret as a number or a date.
        if normalized_col == "s. no.":
            text = str(value).strip()
            return int(text) if text.isdigit() else text

        if normalized_col in {"order no", "certificate no", "customer name",
                              "item", "product type", "manufacturer", "est shipping date"}:
            return str(value).strip()

        if normalized_col == "amount":
            amount = str(value).strip().replace("$", "").replace(",", "")
            try:
                return float(amount)
            except ValueError:
                return str(value).strip()

        # Try numeric coercion
        try:
            v = str(value).strip()
            if re.match(r"^-?\d+$", v):
                # A zero-padded value is an identifier, not a quantity:
                # "00123" is a reference number and int() would silently drop
                # the padding. The built-in layout's identifier columns are
                # named above, but a configured layout can call them anything,
                # so this has to hold for every column.
                if len(v.lstrip("-")) > 1 and v.lstrip("-").startswith("0"):
                    return v
                return int(v)
            if re.match(r"^-?\d+\.\d+$", v):
                return float(v)
        except Exception:
            pass
        return value

    def _apply_cell_format(self, cell: Any, col_name: str) -> None:
        normalized_col = col_name.strip().lower()
        if normalized_col in {"order no", "certificate no"}:
            cell.number_format = "@"
        elif normalized_col == "amount" and isinstance(cell.value, (int, float)):
            cell.number_format = '$#,##0.00'

    def _auto_size_columns(
        self, ws: Any, df: pd.DataFrame, row_offset: int
    ) -> None:
        min_w = self.config.get("min_col_width", 8)
        max_w = self.config.get("max_col_width", 50)

        for col_idx, col_name in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_idx)
            # Measure header
            max_len = len(str(col_name))
            # Sample up to 100 rows for performance. Selected by position, not
            # by label: a promoted header row can produce duplicate column
            # names, and label lookup then returns a DataFrame instead of a
            # Series, so `.str` raised and failed the whole file.
            sample = df.iloc[:100, col_idx - 1].astype(str)
            if not sample.empty:
                max_len = max(max_len, sample.str.len().max())
            ws.column_dimensions[col_letter].width = min(
                max(min_w, max_len + 2), max_w
            )

    def add_text_sheet(self, text: str, sheet_name: str) -> None:
        ws = self.wb.create_sheet(self._unique_sheet_name(sheet_name))
        for line_num, line in enumerate(text.split("\n"), 1):
            ws.cell(row=line_num, column=1, value=line)
        ws.column_dimensions["A"].width = 80

    def add_metadata_sheet(self, meta: Dict) -> None:
        ws = self.wb.create_sheet(self._unique_sheet_name("_Metadata"))
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 50

        header_fill = PatternFill(
            "solid", fgColor=self.config.get("header_color", "1F4E79")
        )
        for row, (key, value) in enumerate(meta.items(), 1):
            k_cell = ws.cell(row=row, column=1, value=str(key))
            v_cell = ws.cell(row=row, column=2, value=str(value))
            k_cell.font = Font(bold=True, color="FFFFFF")
            k_cell.fill = header_fill

    def save(self, output_path: str) -> bool:
        try:
            os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
            self.wb.save(output_path)
            return True
        except Exception as e:
            self.logger.error(f"Failed to save Excel '{output_path}': {e}")
            return False


# ────────────────────────────────────────────────────────────
# PDF PROCESSOR  (orchestrator)
# ────────────────────────────────────────────────────────────

class PDFProcessor:
    """Orchestrates classification → extraction → Excel assembly."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.classifier = PDFClassifier(logger, config.get("pdf_password"))
        self.table_extractor = TableExtractor(config, logger)
        self.text_extractor = TextExtractor(logger, config)
        self.ocr = OCRProcessor(config, logger)

    def extract_datasets(
        self, pdf_path: str, info: Dict[str, Any], preview: bool = False
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, str]]]:
        """Run the full extraction cascade once and return everything it found.

        Returns ``(datasets, text_blocks)``, where a dataset is
        ``{"name", "title", "dataframe"}`` and a text block is
        ``{"name", "text"}`` for a page that yielded no tabular data.

        This is the single source of truth for what a PDF contains; the Excel
        writer and the CSV exporter are both consumers of it. They used to carry
        separate cascades, and the CSV one covered only the table engines - so
        order reports, which are read further down the cascade by the column
        parser, exported nothing at all.
        """
        # Per-document state: this processor is reused across a batch run.
        self.text_extractor.dropped_rows = []
        self.text_extractor.matched_layouts = []
        self.text_extractor.column_formats = {}

        datasets: List[Dict[str, Any]] = []
        text_blocks: List[Dict[str, str]] = []

        # ──────────────── Text pages ────────────────────────────────────────
        if info["text_pages"]:
            tables = self.table_extractor.extract(pdf_path, info["text_pages"])
            combined_tables_df = self._combine_extracted_tables(tables)
            if combined_tables_df is not None and not combined_tables_df.empty:
                datasets.append({
                    "name": "Table_Data",
                    "title": "Extracted Table Data",
                    "dataframe": combined_tables_df,
                })
            else:
                for item in tables:
                    datasets.append({
                        "name": f"P{item['page']}_T{item['table_index']}",
                        "title": f"Page {item['page']} - Table {item['table_index']}",
                        "dataframe": item["dataframe"],
                    })

            # Pages the table engines could not read fall through to the text
            # path, which is where the column parser reads order reports.
            #
            # This used to be all-or-nothing: any table found anywhere skipped
            # the text path for the whole document. One incidental ruled table
            # on page 1 was therefore enough to stop an order report on page 2
            # from ever being read - and to silence its excluded-row accounting
            # along with it.
            pages_with_tables = {item["page"] - 1 for item in tables}
            remaining = [p for p in info["text_pages"] if p not in pages_with_tables]

            if remaining:
                text_data = self.text_extractor.extract(pdf_path, remaining)
                combined_df = self._combine_text_dataframes(text_data)
                if combined_df is not None and not combined_df.empty:
                    datasets.append({
                        "name": "Text_Data",
                        "title": "Extracted Text Data",
                        "dataframe": combined_df,
                        "formats": dict(self.text_extractor.column_formats),
                    })
                else:
                    for item in text_data:
                        name = f"P{item['page']}_Text"
                        df = item.get("dataframe")
                        if df is not None and not df.empty:
                            datasets.append({
                                "name": name,
                                "title": f"Page {item['page']} - Text",
                                "dataframe": df,
                                "formats": dict(self.text_extractor.column_formats),
                            })
                        else:
                            text_blocks.append({"name": name, "text": item["text"]})

        # ──────────────── Scanned pages ──────── OCR ────────────────────────
        if info["scanned_pages"]:
            if not TESSERACT_AVAILABLE:
                self.logger.warning(
                    "Scanned pages found but pytesseract not available. "
                    "Install tesseract-ocr and pytesseract."
                )
            # One open document for the whole OCR loop instead of one per page.
            with _PDFHandles(pdf_path, self.logger, self.config.get("pdf_password")) as ocr_handles:
                for page_num in info["scanned_pages"]:
                    img = self.ocr.page_to_image(pdf_path, page_num, ocr_handles)
                    if img is None:
                        continue
                    df = self.ocr.extract_table_data(img)
                    if df is not None and not df.empty:
                        datasets.append({
                            "name": f"P{page_num+1}_OCR",
                            "title": f"Page {page_num+1} - OCR",
                            "dataframe": df,
                        })
                    else:
                        text = self.ocr.extract_text(img)
                        if text.strip():
                            text_blocks.append({
                                "name": f"P{page_num+1}_OCR_Text",
                                "text": text,
                            })

        if preview:
            for ds in datasets:
                _preview_dataframe(ds["dataframe"], ds["name"])

        return datasets, text_blocks

    def process(
        self, pdf_path: str, output_path: str,
        preview: bool = False
    ) -> Dict:
        result = {
            "input": pdf_path,
            "output": output_path,
            "success": False,
            "pages": 0,
            "tables_found": 0,
            "rows_extracted": 0,
            "sheets_created": 0,
            "unparsed_rows": 0,
            "error": None,
            "duration_sec": 0
        }

        start = datetime.now()
        # Per-document state is reset by extract_datasets(), which owns it.

        if not os.path.exists(pdf_path):
            result["error"] = "File not found"
            return result

        if os.path.getsize(pdf_path) == 0:
            result["error"] = "Empty file"
            return result

        # Checked up front so an encrypted file gets a message naming the fix,
        # rather than an empty extraction that looks like an unreadable PDF.
        access = check_pdf_access(pdf_path, self.config.get("pdf_password"))
        if access == ACCESS_PASSWORD_REQUIRED:
            result["error"] = (
                "Password required. Pass --password, or set "
                f"{PASSWORD_ENV_VAR} in the environment."
            )
            return result
        if access == ACCESS_PASSWORD_INCORRECT:
            result["error"] = "Password incorrect for this PDF"
            return result

        try:
            info = self.classifier.classify(pdf_path)
            result["pages"] = info["page_count"]

            if info["page_count"] <= 0:
                result["error"] = "Unable to read any pages from PDF"
                return result

            builder = ExcelBuilder(self.config, self.logger)

            datasets, text_blocks = self.extract_datasets(
                pdf_path, info, preview=preview
            )

            # Two distinct measures: how many tabular datasets were written,
            # and how many data rows they hold. Conflating them made a 95-row
            # consolidated report display as "95 tables".
            tables_found = len(datasets)
            rows_extracted = sum(len(ds["dataframe"]) for ds in datasets)

            for ds in datasets:
                builder.add_dataframe(ds["dataframe"], ds["name"],
                                      title=ds["title"],
                                      formats=ds.get("formats"))
            for block in text_blocks:
                builder.add_text_sheet(block["text"], block["name"])

            # Last resort: nothing tabular and no text block was captured, so
            # dump the raw page text rather than saving an empty workbook. This
            # runs *before* the metadata sheet is added - adding metadata first
            # made the sheet count non-zero, so this path could never fire.
            if not builder.wb.sheetnames:
                text_data = self.text_extractor.extract(
                    pdf_path, list(range(info["page_count"]))
                )
                for item in text_data:
                    builder.add_text_sheet(item["text"], f"P{item['page']}")

            # ──────────────── Metadata ────────────────────────────────────────────────────────────
            if self.config.get("add_metadata_sheet", True):
                meta = {
                    "Source File": os.path.basename(pdf_path),
                    "Full Path": os.path.abspath(pdf_path),
                    "Processed At": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Page Count": info["page_count"],
                    "PDF Type": info["type"],
                    "Text Pages": len(info["text_pages"]),
                    "Scanned Pages": len(info["scanned_pages"]),
                    "Tables Found": tables_found,
                    "Rows Extracted": rows_extracted,
                    "Unparsed Report Rows": len(self.text_extractor.dropped_rows),
                    # Which layout read the file, so the workbook says what
                    # shape it was understood as rather than leaving it implied.
                    "Report Layout": ", ".join(
                        self.text_extractor.matched_layouts) or "not a known report",
                    "Output File": os.path.abspath(output_path),
                    "Tool": "pdf_to_excel.py"
                }
                builder.add_metadata_sheet(meta)

            # Recorded after every sheet has been added, so the counts describe
            # the workbook that is actually written.
            result["sheets_created"] = len(builder.wb.sheetnames)
            result["tables_found"] = tables_found
            result["rows_extracted"] = rows_extracted
            result["unparsed_rows"] = len(self.text_extractor.dropped_rows)

            saved = builder.save(output_path)
            result["success"] = saved

        except Exception as e:
            self.logger.exception(f"Processing failed for '{pdf_path}': {e}")
            result["error"] = str(e)

        result["duration_sec"] = (datetime.now() - start).total_seconds()
        return result

    def _combine_text_dataframes(
        self, text_data: List[Dict[str, Any]]
    ) -> Optional[pd.DataFrame]:
        frames: List[pd.DataFrame] = []
        base_columns: Optional[List[str]] = None

        for item in text_data:
            df = item.get("dataframe")
            if df is None or df.empty or list(df.columns) == ["Content"]:
                return None

            cols = list(df.columns)
            if base_columns is None:
                base_columns = cols
            elif cols != base_columns:
                return None

            frames.append(df.copy())

        if not frames:
            return None

        return pd.concat(frames, ignore_index=True)

    def _combine_extracted_tables(
        self, tables: List[Dict[str, Any]]
    ) -> Optional[pd.DataFrame]:
        frames: List[pd.DataFrame] = []
        base_columns: Optional[List[str]] = None

        for item in tables:
            df = item.get("dataframe")
            if df is None or df.empty:
                continue

            cols = list(df.columns)
            if base_columns is None:
                base_columns = cols
            elif cols != base_columns:
                return None

            frames.append(df.copy())

        if not frames:
            return None

        return pd.concat(frames, ignore_index=True).drop_duplicates().reset_index(drop=True)


# ────────────────────────────────────────────────────────────
# BATCH PROCESSOR
# ────────────────────────────────────────────────────────────

class BatchProcessor:
    """Process a folder of PDFs with progress tracking and summary report."""

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self.processor = PDFProcessor(config, logger)

    def process_folder(
        self, input_dir: str, output_dir: str,
        preview: bool = False
    ) -> List[Dict]:
        # Deduplicate (a case-insensitive filesystem matches both globs against
        # the same file) then re-sort, so the processing order is deterministic
        # rather than whatever order the intermediate set happened to produce.
        pdfs = sorted({
            p.resolve()
            for p in list(Path(input_dir).glob("**/*.pdf"))
            + list(Path(input_dir).glob("**/*.PDF"))
        })

        if not pdfs:
            _print("No PDF files found in: " + input_dir)
            return []

        os.makedirs(output_dir, exist_ok=True)
        results = []
        taken: Dict[str, Path] = {}

        if RICH_AVAILABLE:
            with Progress(
                SpinnerColumn(),
                TextColumn("[bold blue]{task.description}"),
                BarColumn(),
                TaskProgressColumn(),
                TimeElapsedColumn(),
            ) as progress:
                task = progress.add_task("Processing PDFs...", total=len(pdfs))
                for pdf in pdfs:
                    out = self._reserve_output_path(pdf, output_dir, taken)
                    progress.update(task, description=f"[bold blue]{pdf.name}")
                    r = self.processor.process(str(pdf), out, preview=preview)
                    results.append(r)
                    progress.advance(task)
        else:
            for i, pdf in enumerate(pdfs, 1):
                print(f"[{i}/{len(pdfs)}] Processing: {pdf.name}")
                out = self._reserve_output_path(pdf, output_dir, taken)
                r = self.processor.process(str(pdf), out, preview=preview)
                results.append(r)
                status = "OK" if r["success"] else f"FAIL: {r['error']}"
                print(f"  -> {status}")

        if self.config.get("batch_summary_report", True):
            self._write_summary_report(results, output_dir)

        return results

    def _reserve_output_path(
        self, pdf: Path, output_dir: str, taken: Dict[str, Path]
    ) -> str:
        """Claim an output name no other file in this run is using.

        Batch mode recurses subdirectories, so two different PDFs can share a
        stem - "jan/orders.pdf" and "feb/orders.pdf". Naming the output from the
        stem alone made the second overwrite the first while the batch summary
        reported both as succeeded: an entire converted workbook lost with no
        indication that anything had happened.

        Uniqueness is scoped to the run. Re-running a batch still refreshes the
        previous run's workbooks, which is the expected behaviour; only a clash
        *within* one run is a defect.
        """
        base = f"{pdf.stem}_{datetime.now().strftime('%Y%m%d')}"
        name, counter = base, 1
        while name in taken:
            counter += 1
            name = f"{base}_{counter}"

        if counter > 1:
            self.logger.warning(
                f"Output name '{base}.xlsx' was already claimed by "
                f"'{taken[base]}'; writing '{pdf}' to '{name}.xlsx' instead."
            )

        taken[name] = pdf
        return os.path.join(output_dir, f"{name}.xlsx")

    def _write_summary_report(self, results: List[Dict], output_dir: str) -> None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = os.path.join(output_dir, f"_batch_summary_{ts}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"

        headers = [
            "Input File", "Output File", "Success", "Pages",
            "Tables Found", "Rows Extracted", "Sheets Created",
            "Unparsed Rows", "Duration (s)", "Error"
        ]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        # Column positions derived from `headers` so they cannot drift apart.
        success_col = headers.index("Success") + 1
        unparsed_col = headers.index("Unparsed Rows") + 1
        duration_col = headers.index("Duration (s)") + 1

        successes = 0
        for r_idx, r in enumerate(results, 2):
            unparsed = r.get("unparsed_rows", 0)
            values = [
                os.path.basename(r["input"]),
                os.path.basename(r.get("output", "")),
                "Yes" if r["success"] else "No",
                r["pages"],
                r["tables_found"],
                r.get("rows_extracted", 0),
                r["sheets_created"],
                unparsed,
                round(r["duration_sec"], 2),
                r.get("error") or "",
            ]
            for c_idx, value in enumerate(values, 1):
                ws.cell(r_idx, c_idx, value)
            if r["success"]:
                successes += 1
                ws.cell(r_idx, success_col).fill = PatternFill("solid", fgColor="C6EFCE")
            else:
                ws.cell(r_idx, success_col).fill = PatternFill("solid", fgColor="FFC7CE")
            # Flag files where report rows were silently unmatched.
            if unparsed:
                ws.cell(r_idx, unparsed_col).fill = PatternFill("solid", fgColor="FFEB9C")
                ws.cell(r_idx, unparsed_col).font = Font(bold=True)

        # Auto-size
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        # Stats row
        last = len(results) + 2
        ws.cell(last, 1, "TOTAL").font = Font(bold=True)
        ws.cell(last, success_col, f"{successes}/{len(results)} succeeded").font = Font(bold=True)
        total_rows = sum(r.get("rows_extracted", 0) for r in results)
        ws.cell(last, headers.index("Rows Extracted") + 1, total_rows).font = Font(bold=True)
        total_unparsed = sum(r.get("unparsed_rows", 0) for r in results)
        ws.cell(last, unparsed_col, total_unparsed).font = Font(bold=True)
        total_dur = sum(r["duration_sec"] for r in results)
        ws.cell(last, duration_col, round(total_dur, 2)).font = Font(bold=True)

        wb.save(report_path)
        _print(f"Batch summary: {report_path}")
        if total_unparsed:
            _print(
                f"[yellow]Warning: {total_unparsed} report row(s) across the batch did not "
                f"match the expected format and were excluded.[/yellow]"
                if RICH_AVAILABLE else
                f"Warning: {total_unparsed} report row(s) across the batch did not "
                f"match the expected format and were excluded."
            )


# ────────────────────────────────────────────────────────────
# DOCUMENT → PDF  (headless LibreOffice)
# ────────────────────────────────────────────────────────────

# A fixed list rather than "try anything". An unsupported file otherwise fails
# deep inside soffice with an opaque message, long after the caller could have
# been told plainly that the format is not handled.
DOC_TO_PDF_EXTENSIONS = {
    ".doc", ".docx", ".odt", ".rtf", ".txt",
    ".xls", ".xlsx", ".ods", ".csv",
    ".ppt", ".pptx", ".odp",
    ".html", ".htm",
}

# Default install locations, checked only after PATH.
_LIBREOFFICE_HINTS = (
    r"C:\Program Files\LibreOffice\program\soffice.exe",
    r"C:\Program Files (x86)\LibreOffice\program\soffice.exe",
    "/Applications/LibreOffice.app/Contents/MacOS/soffice",
    "/usr/bin/soffice",
    "/usr/bin/libreoffice",
    "/usr/local/bin/soffice",
    "/snap/bin/libreoffice",
)

_UNRESOLVED = object()   # binary lookup not attempted yet


class DocumentToPDFConverter:
    """Convert office documents to PDF with a headless LibreOffice.

    LibreOffice is system software pip cannot install, so it follows the same
    rule as Tesseract, Ghostscript and Java: located at runtime, reported by
    --check-deps, and when it is absent the caller gets a clear message instead
    of a crash or an empty output directory.

    This is the one direction the tool converts *into* PDF. Everything else
    here reads a PDF; this produces one, which is useful when a report arrives
    as .docx or .xlsx and has to become a PDF before anything else can run.
    """

    def __init__(self, config: dict, logger: logging.Logger):
        self.config = config
        self.logger = logger
        self._binary: Any = _UNRESOLVED

    # ── locating LibreOffice ─────────────────────────────────
    def binary(self) -> Optional[str]:
        if self._binary is _UNRESOLVED:
            self._binary = self._find_binary()
        return self._binary

    def _find_binary(self) -> Optional[str]:
        configured = self.config.get("libreoffice_path")
        if configured:
            if os.path.isfile(configured):
                return configured
            # Say so rather than silently searching elsewhere: a wrong path in
            # config is a mistake the user wants to hear about.
            self.logger.warning(
                f"libreoffice_path is set to '{configured}' but no file is there. "
                f"Looking for LibreOffice on PATH instead."
            )

        for name in ("soffice", "libreoffice"):
            found = shutil.which(name)
            if found:
                return found

        for path in _LIBREOFFICE_HINTS:
            if os.path.isfile(path):
                return path

        return None

    def available(self) -> bool:
        return self.binary() is not None

    # ── command construction ─────────────────────────────────
    @staticmethod
    def _command(
        binary: str, src: str, out_dir: str, profile_dir: str
    ) -> List[str]:
        """Build the soffice invocation.

        `-env:UserInstallation` gives this run a private profile directory.
        Without it, starting soffice while another instance holds the default
        profile makes the new process hand off the job and exit 0 immediately -
        having written nothing. That failure is silent, produces a success exit
        code, and is why batch runs and CI need the isolation.
        """
        return [
            binary,
            f"-env:UserInstallation={Path(profile_dir).as_uri()}",
            "--headless",
            "--norestore",
            "--invisible",
            "--convert-to", "pdf",
            "--outdir", out_dir,
            src,
        ]

    # ── conversion ───────────────────────────────────────────
    def convert(
        self, input_path: str, output_dir: str, timeout: int = 180,
        output_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Convert one document to PDF. Returns a result dict; never raises.

        `output_name` (without extension) overrides the name taken from the
        source stem, which is how convert_folder keeps two same-named documents
        from landing on top of each other.
        """
        result: Dict[str, Any] = {
            "input": input_path,
            "output": None,
            "success": False,
            "error": None,
            "duration_sec": 0.0,
        }
        start = datetime.now()

        try:
            ext = Path(input_path).suffix.lower()

            if not os.path.exists(input_path):
                result["error"] = "File not found"
                return result
            if ext == ".pdf":
                result["error"] = "Already a PDF"
                return result
            if ext not in DOC_TO_PDF_EXTENSIONS:
                result["error"] = (
                    f"Unsupported format '{ext}'. Supported: "
                    + ", ".join(sorted(DOC_TO_PDF_EXTENSIONS))
                )
                return result

            binary = self.binary()
            if binary is None:
                result["error"] = (
                    "LibreOffice not found. Install it and put 'soffice' on PATH, "
                    "or set 'libreoffice_path' in config.json."
                )
                return result

            os.makedirs(output_dir, exist_ok=True)
            final_path = os.path.join(
                output_dir, (output_name or Path(input_path).stem) + ".pdf"
            )

            # soffice names its output after the source stem and offers no way
            # to override it, so it writes into a staging directory and the
            # result is moved to the name the caller reserved. That is what
            # lets two same-named sources coexist in one output folder.
            with tempfile.TemporaryDirectory(prefix="lo_") as work:
                profile = os.path.join(work, "profile")
                staged_dir = os.path.join(work, "out")
                os.makedirs(staged_dir)

                cmd = self._command(binary, os.path.abspath(input_path),
                                    staged_dir, profile)
                self.logger.debug("LibreOffice: " + " ".join(cmd))
                try:
                    proc = subprocess.run(
                        cmd, capture_output=True, text=True, timeout=timeout
                    )
                except subprocess.TimeoutExpired:
                    result["error"] = f"LibreOffice timed out after {timeout}s"
                    return result
                except OSError as e:
                    result["error"] = f"Could not run LibreOffice: {e}"
                    return result

                staged = os.path.join(staged_dir, Path(input_path).stem + ".pdf")

                # The exit code alone is not trustworthy - soffice returns 0 in
                # cases where it wrote nothing at all - so the output file is
                # the thing actually checked.
                if not os.path.isfile(staged):
                    detail = (proc.stderr or proc.stdout or "").strip().splitlines()
                    result["error"] = (
                        f"LibreOffice produced no PDF (exit {proc.returncode})"
                        + (f": {detail[-1][:200]}" if detail else "")
                    )
                    return result

                shutil.move(staged, final_path)

            result["output"] = final_path
            result["success"] = True
            self.logger.info(f"Converted '{input_path}' -> '{final_path}'")
            return result

        except Exception as e:                       # pragma: no cover - defensive
            self.logger.exception(f"Document conversion failed for '{input_path}': {e}")
            result["error"] = str(e)
            return result
        finally:
            result["duration_sec"] = (datetime.now() - start).total_seconds()

    def _reserve_name(self, src: Path, taken: Dict[str, Path]) -> str:
        """Claim an output name no other document in this run is using.

        The folder walk is recursive, so 'jan/report.txt' and 'feb/report.html'
        both want report.pdf. Letting them share it meant the second silently
        overwrote the first while both were reported as converted - the same
        loss BatchProcessor guards against on the Excel side.
        """
        base = src.stem
        name, counter = base, 1
        while name in taken:
            counter += 1
            name = f"{base}_{counter}"

        if counter > 1:
            self.logger.warning(
                f"Output name '{base}.pdf' was already claimed by "
                f"'{taken[base]}'; writing '{src}' to '{name}.pdf' instead."
            )

        taken[name] = src
        return name

    def convert_folder(
        self, input_dir: str, output_dir: str
    ) -> List[Dict[str, Any]]:
        """Convert every supported document in a folder, recursively."""
        files = sorted({
            p.resolve()
            for p in Path(input_dir).glob("**/*")
            if p.is_file() and p.suffix.lower() in DOC_TO_PDF_EXTENSIONS
        })

        if not files:
            _print(f"No convertible documents found in: {input_dir}")
            return []

        results = []
        taken: Dict[str, Path] = {}
        for i, path in enumerate(files, 1):
            print(f"[{i}/{len(files)}] Converting: {path.name}")
            r = self.convert(str(path), output_dir,
                             output_name=self._reserve_name(path, taken))
            results.append(r)
            # ASCII arrow on purpose: this goes through plain print(), which
            # encodes with the console codepage. A "→" raises UnicodeEncodeError
            # on a cp1252 console - including any Windows run redirected to a
            # file - and would take down the whole batch mid-way.
            print(f"  -> {'OK' if r['success'] else 'FAIL: ' + str(r['error'])}")
        return results


# ────────────────────────────────────────────────────────────
# UTILITIES
# ────────────────────────────────────────────────────────────

def _print(msg: str) -> None:
    if RICH_AVAILABLE:
        console.print(msg)
    else:
        print(msg)


def _preview_dataframe(df: pd.DataFrame, label: str, n: int = 10) -> None:
    _print(f"\n[bold]Preview: {label}[/bold]" if RICH_AVAILABLE else f"\nPreview: {label}")
    if RICH_AVAILABLE:
        t = Table(show_header=True, header_style="bold blue")
        cols = list(df.columns)[:8]
        for c in cols:
            t.add_column(str(c), max_width=20)
        for _, row in df.head(n).iterrows():
            t.add_row(*[str(row[c])[:20] for c in cols])
        console.print(t)
    else:
        print(df.head(n).to_string())


def libreoffice_present(config: Optional[dict] = None) -> bool:
    """Whether a headless LibreOffice can be found, without running it."""
    configured = (config or {}).get("libreoffice_path")
    if configured and os.path.isfile(configured):
        return True
    if shutil.which("soffice") or shutil.which("libreoffice"):
        return True
    return any(os.path.isfile(p) for p in _LIBREOFFICE_HINTS)


def check_dependencies(config: Optional[dict] = None) -> Dict[str, bool]:
    deps = {
        "pdfplumber": PDFPLUMBER_AVAILABLE,
        "PyMuPDF (fitz)": PYMUPDF_AVAILABLE,
        "camelot-py": CAMELOT_AVAILABLE,
        "tabula-py": TABULA_AVAILABLE,
        "pytesseract": TESSERACT_AVAILABLE,
        "Pillow": PILLOW_AVAILABLE,
        "rich": RICH_AVAILABLE,
        "pandas": True,
        "openpyxl": True,
        "numpy": True,
        # System software, not a Python package - only needed for --to-pdf.
        "LibreOffice (--to-pdf)": libreoffice_present(config),
    }
    return deps


def print_dependency_status(config: Optional[dict] = None) -> None:
    deps = check_dependencies(config)
    if RICH_AVAILABLE:
        t = Table(title="Dependency Status", show_header=True)
        t.add_column("Library", style="cyan")
        t.add_column("Available", style="bold")
        for name, ok in deps.items():
            t.add_row(name, "[green]YES[/green]" if ok else "[red]NO[/red]")
        console.print(t)
    else:
        print("\nDependency Status:")
        for name, ok in deps.items():
            print(f"  {'[OK]' if ok else '[MISSING]':8} {name}")


# ────────────────────────────────────────────────────────────
# CSV EXPORT
# ────────────────────────────────────────────────────────────

def export_to_csv(
    pdf_path: str, output_dir: str, config: dict, logger: logging.Logger
) -> Dict[str, Any]:
    """Extract every dataset in a PDF and save each as a separate CSV.

    Returns ``{"files": [paths], "unparsed_rows": int}``.

    This runs the same cascade as the Excel path via
    `PDFProcessor.extract_datasets`. It previously called `TableExtractor`
    directly, which meant it saw only what the table engines found: order
    reports produce no ruled grid and are read further down the cascade by the
    column parser, so exporting one as CSV silently produced no files at all.
    Sharing the cascade also means CSV runs now cover scanned pages via OCR,
    and can report excluded rows - which the caller must surface, because a
    partial export that looks complete is the one outcome this tool rejects.
    """
    processor = PDFProcessor(config, logger)
    info = processor.classifier.classify(pdf_path)
    datasets, _text_blocks = processor.extract_datasets(pdf_path, info)

    os.makedirs(output_dir, exist_ok=True)
    stem = Path(pdf_path).stem

    saved = []
    for ds in datasets:
        # Named for the dataset, so a CSV file and the equivalent Excel sheet
        # carry the same name.
        fpath = os.path.join(output_dir, f"{stem}_{ds['name']}.csv")
        ds["dataframe"].to_csv(fpath, index=False, encoding="utf-8-sig")
        saved.append(fpath)

    return {
        "files": saved,
        "unparsed_rows": len(processor.text_extractor.dropped_rows),
    }


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        prog="pdf_to_excel",
        description="Convert PDF files to Excel with AI/OCR capabilities.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Single file
  python pdf_to_excel.py report.pdf

  # Single file, custom output
  python pdf_to_excel.py invoice.pdf -o results/invoice.xlsx

  # Batch folder
  python pdf_to_excel.py -d ./pdfs/ -o ./output/

  # With preview and CSV export
  python pdf_to_excel.py report.pdf --preview --format csv

  # Use specific config
  python pdf_to_excel.py report.pdf --config my_config.json

  # Convert documents INTO PDF (needs LibreOffice)
  python pdf_to_excel.py report.docx --to-pdf -o ./pdfs/
  python pdf_to_excel.py -d ./documents/ --to-pdf -o ./pdfs/

  # Launch the web UI (it lives in app.py, not here)
  streamlit run app.py

  # Check dependencies
  python pdf_to_excel.py --check-deps
        """
    )
    p.add_argument("input", nargs="?", help="PDF file or (with -d) unused")
    p.add_argument("-d", "--dir", help="Input directory for batch processing")
    p.add_argument("-o", "--output", help="Output .xlsx file or output directory")
    p.add_argument("--format", choices=["excel", "csv"], default="excel",
                   help="Output format (default: excel)")
    p.add_argument("--config", default="config.json",
                   help="Config file path (default: config.json)")
    p.add_argument("--preview", action="store_true",
                   help="Print table preview before saving")
    p.add_argument("--verbose", action="store_true",
                   help="Verbose logging")
    p.add_argument("--method",
                   choices=["auto", "pdfplumber", "camelot", "tabula"],
                   default=None, help="Force table extraction method")
    p.add_argument("--ocr-lang", default=None,
                   help="Tesseract language code (default: eng)")
    p.add_argument("--no-meta", action="store_true",
                   help="Omit metadata sheet")
    p.add_argument("--keep-empty", action="store_true",
                   help="Keep empty rows and columns")
    p.add_argument("--password", default=None,
                   help="Password for an encrypted PDF. Prefer the "
                        f"{PASSWORD_ENV_VAR} environment variable: an argument "
                        "is visible in shell history and process listings")
    p.add_argument("--to-pdf", action="store_true",
                   help="Convert documents INTO PDF via headless LibreOffice "
                        "(.docx/.xlsx/.pptx/.odt/.rtf/.txt/.html and more), "
                        "instead of reading a PDF. Works with -d for a folder")
    p.add_argument("--gui", action="store_true",
                   help="Print the command that launches the web UI, and exit "
                        "(the UI lives in app.py; this does not start it)")
    p.add_argument("--check-deps", action="store_true",
                   help="Print dependency status and exit")
    return p


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.check_deps:
        # Config is read first so a configured libreoffice_path is honoured by
        # the status report rather than contradicted by it.
        print_dependency_status(load_config(args.config))
        return 0

    if args.gui:
        print("Launch GUI with: streamlit run app.py")
        print("Or double-click convert.bat")
        return 0

    cfg = load_config(args.config)
    logger = setup_logging(cfg.get("log_dir", "logs"), verbose=args.verbose)

    # Runtime only, never persisted: see PASSWORD_ENV_VAR.
    cfg["pdf_password"] = args.password or os.environ.get(PASSWORD_ENV_VAR)

    # Apply CLI overrides
    if args.method:
        cfg["table_extraction_method"] = args.method
    if args.ocr_lang:
        cfg["ocr_language"] = args.ocr_lang
    if args.no_meta:
        cfg["add_metadata_sheet"] = False
    if args.keep_empty:
        cfg["remove_empty_rows"] = False
        cfg["remove_empty_cols"] = False

    # ──────────────── Document → PDF ────────────────────────────────────────
    if args.to_pdf:
        converter = DocumentToPDFConverter(cfg, logger)
        out_dir = args.output or cfg.get("output_dir", "output")

        if not converter.available():
            _print(
                "[red]LibreOffice not found.[/red] --to-pdf needs it. Install "
                "LibreOffice and put 'soffice' on PATH, or set "
                "'libreoffice_path' in config.json."
                if RICH_AVAILABLE else
                "LibreOffice not found. --to-pdf needs it. Install LibreOffice "
                "and put 'soffice' on PATH, or set 'libreoffice_path' in config.json."
            )
            return 1

        if args.dir:
            results = converter.convert_folder(args.dir, out_dir)
        elif args.input:
            results = [converter.convert(args.input, out_dir)]
        else:
            _print("--to-pdf needs an input file, or -d for a folder.")
            return 1

        if not results:
            return 1

        ok = sum(1 for r in results if r["success"])
        fail = len(results) - ok
        _print(f"\nConverted {ok} of {len(results)} to PDF in: {out_dir}")
        for r in results:
            if not r["success"]:
                msg = f"Failed: {os.path.basename(r['input'])} - {r['error']}"
                _print(f"[red]{msg}[/red]" if RICH_AVAILABLE else msg)
        return 0 if fail == 0 else 1

    # ──────────────── Batch mode ────────────────────────────────────────────────────────────
    if args.dir:
        output_dir = args.output or cfg.get("output_dir", "output")
        batch = BatchProcessor(cfg, logger)
        results = batch.process_folder(args.dir, output_dir, preview=args.preview)

        ok = sum(1 for r in results if r["success"])
        fail = len(results) - ok
        _print(f"\nBatch complete: {ok} succeeded, {fail} failed out of {len(results)}")
        return 0 if fail == 0 else 1

    # ──────────────── Single file mode ────────────────────────────────────────────────────────────
    if not args.input:
        parser.print_help()
        return 1

    pdf_path = args.input
    if not os.path.exists(pdf_path):
        _print(f"[red]Error: File not found: {pdf_path}[/red]" if RICH_AVAILABLE
               else f"Error: File not found: {pdf_path}")
        return 1

    stem = Path(pdf_path).stem
    ts = datetime.now().strftime("%Y%m%d")

    if args.format == "csv":
        output_dir = args.output or os.path.join(cfg.get("output_dir", "output"), stem)
        export = export_to_csv(pdf_path, output_dir, cfg, logger)
        saved = export["files"]
        if saved:
            _print(f"Exported {len(saved)} CSV file(s) to: {output_dir}")
        else:
            _print("No data found to export as CSV.")
        if export["unparsed_rows"]:
            warn = (
                f"Warning: {export['unparsed_rows']} report row(s) did not match the "
                f"expected format and were excluded from the export. "
                f"The offending lines are listed in the log file."
            )
            _print(f"[yellow]{warn}[/yellow]" if RICH_AVAILABLE else warn)
        return 0

    # No interactive password prompt, deliberately. getpass on Windows reads
    # the console directly rather than stdin, so `isatty()` does not reliably
    # predict whether anyone is there to answer - a scheduled or piped run can
    # sit waiting forever. Hanging is a worse failure than a clear message, so
    # the CLI states what to supply and exits; the web UI is the interactive
    # surface and asks there.
    output_path = args.output or os.path.join(
        cfg.get("output_dir", "output"), f"{stem}_{ts}.xlsx"
    )

    if RICH_AVAILABLE:
        console.print(Panel(
            f"[bold]Input:[/bold]  {pdf_path}\n"
            f"[bold]Output:[/bold] {output_path}",
            title="PDF to Excel", border_style="blue"
        ))

    processor = PDFProcessor(cfg, logger)
    result = processor.process(pdf_path, output_path, preview=args.preview)

    if result["success"]:
        msg = (
            f"Done in {result['duration_sec']:.1f}s | "
            f"{result['pages']} pages | "
            f"{result['tables_found']} tables | "
            f"{result['rows_extracted']} rows | "
            f"{result['sheets_created']} sheets -> {output_path}"
        )
        _print(f"[green]{msg}[/green]" if RICH_AVAILABLE else msg)
        if result["unparsed_rows"]:
            warn = (
                f"Warning: {result['unparsed_rows']} report row(s) did not match the "
                f"expected format and were excluded from the output. "
                f"The offending lines are listed in the log file."
            )
            _print(f"[yellow]{warn}[/yellow]" if RICH_AVAILABLE else warn)
        return 0
    else:
        err_msg = f"Failed: {result['error']}"
        _print(f"[red]{err_msg}[/red]" if RICH_AVAILABLE else err_msg)
        return 1


if __name__ == "__main__":
    sys.exit(main())

