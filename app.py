#!/usr/bin/env python3
"""
PDF to Excel Converter — web interface.
Run: streamlit run app.py   (or double-click convert.bat)

DIRECTION CONTRACT
THESIS: A working ledger, not a file-upload page. It refuses the centred
  dropzone-and-blue-button SaaS uploader by making the extracted data and its
  reconciliation the screen; upload is a small entry step, not the subject.
OWN-WORLD: Cool-grey neutrals on light ground, 1px hairlines instead of
  shadows, native Segoe/Cascadia stacks, tabular numerals on every figure, an
  ink-black primary action (never SaaS blue), and one reserved amber that only
  ever means "rows excluded from the output".
STORY: The operator drops an invoice, sees what the document is, converts, then
  reads a reconciliation against the invoice's own stated total before
  downloading — so downloading confirms something already visibly true.
  Dropping a non-PDF instead routes to document→PDF conversion. The file
  decides; there is no mode switch, because which flow applies is a fact
  already on disk rather than a question worth asking.
FIRST VIEWPORT: Wordmark over a hairline rule; page title left; one bordered
  dropzone panel filling the working column with format guidance beneath. No
  hero, no cards, no metrics until there is a document to describe.
FORM: Enterprise console canon (standing exit), chosen over rolled direction 7
  "bourse board"; staging: rehearsed preview. Seed 7495fcd5.
"""

import os
import re
import sys
import shutil
import tempfile
import time
from datetime import datetime

import streamlit as st
import pandas as pd
from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from pathlib import Path  # noqa: E402
from pdf_to_excel import (  # noqa: E402
    load_config, setup_logging, PDFProcessor,
    PDFClassifier, export_to_csv,
    DocumentToPDFConverter, DOC_TO_PDF_EXTENSIONS, libreoffice_present,
    check_pdf_access, ACCESS_OK, ACCESS_PASSWORD_REQUIRED,
)

# The uploader decides the flow. Dropping a PDF extracts from it; dropping a
# document converts it into one. No mode switch: the file says which it is, so
# asking the operator to declare it first would be chrome over a fact already
# on disk (product principle 4, the document is the authority).
UPLOAD_TYPES = ["pdf"] + sorted(e.lstrip(".") for e in DOC_TO_PDF_EXTENSIONS)

st.set_page_config(
    page_title="PDF to Excel Converter",
    page_icon="◆",
    layout="centered",
)

# ─────────────────────────────────────────────────────────────────────────────
# DESIGN TOKENS  (see DESIGN.md)
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
:root{
  --bg:#fbfbfa; --surface:#fff; --surface-2:#f6f6f5;
  --border:#e6e6e3; --border-strong:#d4d4d0;
  /* text-3 carries 11px labels and table headers, so it must clear 4.5:1
     on both --surface and --surface-2, not merely look light. */
  --text:#18181b; --text-2:#52525b; --text-3:#6b6b74;
  --ink:#18181b; --ink-fg:#fff;
  --held:#b45309; --held-bg:#fef6ec; --held-br:#f0d9b5;
  --ok:#15803d;   --ok-bg:#f0f8f2;   --ok-br:#c9e4d2;
  --fail:#b91c1c; --fail-bg:#fdf2f2; --fail-br:#f2cccc;
  --r-ctl:6px; --r-panel:10px;
  --dur-fast:120ms; --dur:180ms; --ease:cubic-bezier(.2,0,0,1);
  --font:"Segoe UI Variable Text","Segoe UI",-apple-system,BlinkMacSystemFont,system-ui,sans-serif;
  --mono:"Cascadia Mono",Consolas,ui-monospace,"SF Mono",monospace;
}
@media (prefers-color-scheme: dark){
  :root{
    --bg:#0c0c0d; --surface:#141416; --surface-2:#1b1b1e;
    --border:#26262a; --border-strong:#35353a;
    --text:#f4f4f5; --text-2:#a1a1aa; --text-3:#9a9aa4;
    --ink:#fafafa; --ink-fg:#0c0c0d;
    --held:#f0b357; --held-bg:#231a0d; --held-br:#4a3617;
    --ok:#5dc98a;   --ok-bg:#0f1d15;   --ok-br:#1f4030;
    --fail:#f08a8a; --fail-bg:#231314; --fail-br:#4a2224;
  }
}

/* ── Streamlit chrome ─────────────────────────────────────────────── */
#MainMenu, footer, header, [data-testid="stToolbar"],
[data-testid="stDecoration"], [data-testid="stStatusWidget"]{display:none!important;}
.stApp{background:var(--bg);}
html,body,[class*="css"],.stApp{
  font-family:var(--font); color:var(--text);
  -webkit-font-smoothing:antialiased;
}
.block-container{max-width:940px; padding:28px 24px 96px!important;}
/* Vertical rhythm between blocks.
   Streamlit ships `margin-bottom:-16px` on every stMarkdownContainer and
   relies on a default 1rem block gap to net out ~0. Our custom markup made
   that negative margin visible as an overlap: each block's content was pulled
   up over the block below it (worst at the expander and download button).
   Fix at the source — zero that negative margin — then set our own positive
   gap, which cannot collapse and so guarantees clearance everywhere. */
[data-testid="stMarkdownContainer"]{margin-bottom:0!important;}
[data-testid="stMain"] [data-testid="stVerticalBlock"]{gap:14px!important;}
hr{display:none;}

/* ── Masthead ─────────────────────────────────────────────────────── */
.app-mast{
  display:flex; align-items:center; justify-content:space-between;
  gap:16px; padding-bottom:14px; margin-bottom:28px;
  border-bottom:1px solid var(--border);
}
.app-brand{display:flex; align-items:center; gap:11px;}
.app-brand .mark{
  width:28px; height:28px; display:grid; place-items:center;
  color:var(--text); font-size:18px; line-height:1;
}
.app-brand .name{
  font-size:17px; font-weight:600; letter-spacing:-.005em; color:var(--text);
}
.app-env{
  font-family:var(--mono); font-size:11px; color:var(--text-3);
  letter-spacing:.02em;
}

/* ── Type ─────────────────────────────────────────────────────────── */
.app-h1{font-size:28px; line-height:1.2; letter-spacing:-.02em; font-weight:600; margin:0 0 6px;}
.app-sub{font-size:14px; line-height:1.55; color:var(--text-2); margin:0 0 26px; max-width:62ch;}
.app-h2{font-size:18px; line-height:1.35; letter-spacing:-.011em; font-weight:600; margin:20px 0 10px;}
.app-label{
  font-size:11px; font-weight:550; letter-spacing:.04em; text-transform:uppercase;
  color:var(--text-3); margin:0;
}
.app-help{font-size:13px; line-height:1.5; color:var(--text-2); margin:8px 0 0;}
.num{font-variant-numeric:tabular-nums;}
.mono{font-family:var(--mono); font-variant-numeric:tabular-nums;}

/* ── Panel ────────────────────────────────────────────────────────── */
.app-panel{
  background:var(--surface); border:1px solid var(--border);
  border-radius:var(--r-panel); padding:18px 20px;
}
.app-panel + .app-panel{margin-top:12px;}

/* ── Metric row ───────────────────────────────────────────────────── */
.app-metrics{
  display:grid; grid-template-columns:repeat(4,1fr);
  background:var(--surface); border:1px solid var(--border);
  border-radius:var(--r-panel); overflow:hidden;
}
.app-metrics > div{padding:14px 18px; border-left:1px solid var(--border);}
.app-metrics > div:first-child{border-left:0;}
.app-metric-v{
  font-size:24px; font-weight:600; letter-spacing:-.02em; line-height:1.15;
  font-variant-numeric:tabular-nums; margin-top:5px; color:var(--text);
}
@media (max-width:720px){ .app-metrics{grid-template-columns:repeat(2,1fr);}
  .app-metrics > div:nth-child(3){border-left:0;}
  .app-metrics > div:nth-child(n+3){border-top:1px solid var(--border);} }

/* ── Status pills ─────────────────────────────────────────────────── */
.pill{
  display:inline-flex; align-items:center; gap:6px;
  font-size:11px; font-weight:550; letter-spacing:.04em; text-transform:uppercase;
  padding:3px 9px; border-radius:999px; border:1px solid;
}
.pill .dot{width:5px; height:5px; border-radius:50%; background:currentColor;}
.pill.ok{color:var(--ok); background:var(--ok-bg); border-color:var(--ok-br);}
.pill.held{color:var(--held); background:var(--held-bg); border-color:var(--held-br);}
.pill.fail{color:var(--fail); background:var(--fail-bg); border-color:var(--fail-br);}

/* ── Notices ──────────────────────────────────────────────────────── */
.notice{
  border:1px solid; border-radius:var(--r-panel);
  padding:14px 16px; margin-top:12px; font-size:13.5px; line-height:1.55;
}
.notice .t{font-weight:600; display:block; margin-bottom:3px;}
.notice.held{background:var(--held-bg); border-color:var(--held-br); color:var(--held);}
.notice.fail{background:var(--fail-bg); border-color:var(--fail-br); color:var(--fail);}
.notice.held .b, .notice.fail .b{color:var(--text-2); display:block; margin-top:5px;}

/* ── Reconciliation ───────────────────────────────────────────────── */
.recon{display:grid; grid-template-columns:1fr auto auto; gap:2px 24px; align-items:baseline;}
.recon .r{display:contents;}
.recon .k{font-size:13px; color:var(--text-2); padding:5px 0;}
.recon .v{
  font-family:var(--mono); font-size:13px; font-variant-numeric:tabular-nums;
  text-align:right; padding:5px 0; color:var(--text);
}
.recon .sep{grid-column:1/-1; height:1px; background:var(--border); margin:5px 0;}
.recon .k.total,.recon .v.total{font-weight:600; color:var(--text);}

/* Reconciliation verdict — the product's key trust moment, so it leads the
   panel at metric scale rather than whispering in the ledger rows. */
.verdict{
  display:flex; align-items:baseline; gap:14px; flex-wrap:wrap;
  padding-bottom:14px; margin-bottom:14px; border-bottom:1px solid var(--border);
}
.verdict-fig{
  font-size:30px; font-weight:600; letter-spacing:-.02em; line-height:1;
  font-variant-numeric:tabular-nums;
}
.verdict-cap{font-size:13px; color:var(--text-2);}
.verdict.ok .verdict-fig{color:var(--ok);}
.verdict.held .verdict-fig{color:var(--held);}

/* ── Ledger table ─────────────────────────────────────────────────── */
.ledger-wrap{
  border:1px solid var(--border); border-radius:var(--r-panel);
  overflow:hidden; background:var(--surface);
}
.ledger-scroll{overflow-x:auto;}
table.ledger{border-collapse:collapse; width:100%; font-size:13px;}
table.ledger th{
  background:var(--surface-2); color:var(--text-3);
  font-size:11px; font-weight:550; letter-spacing:.04em; text-transform:uppercase;
  text-align:left; padding:9px 12px; white-space:nowrap;
  border-bottom:1px solid var(--border);
}
table.ledger td{
  padding:9px 12px; border-bottom:1px solid var(--border);
  color:var(--text); white-space:nowrap;
}
table.ledger tr:last-child td{border-bottom:0;}
table.ledger tbody tr{transition:background var(--dur-fast) var(--ease);}
table.ledger tbody tr:hover{background:var(--surface-2);}
table.ledger td.n,table.ledger th.n{text-align:right; font-variant-numeric:tabular-nums;}
table.ledger td.m{font-family:var(--mono); font-variant-numeric:tabular-nums;}
.ledger-foot{
  padding:8px 12px; border-top:1px solid var(--border);
  background:var(--surface-2); font-size:12px; color:var(--text-3);
}

/* ── Buttons ──────────────────────────────────────────────────────── */
.stButton > button, .stDownloadButton > button{
  width:100%; border-radius:var(--r-ctl); font-family:var(--font);
  font-size:14px; font-weight:550; padding:9px 16px; min-height:38px;
  transition:background var(--dur-fast) var(--ease),
             border-color var(--dur-fast) var(--ease),
             opacity var(--dur-fast) var(--ease);
  box-shadow:none!important;
}
.stButton > button[kind="primary"],
.stButton > button[data-testid="stBaseButton-primary"]{
  background:var(--ink); color:var(--ink-fg); border:1px solid var(--ink);
}
.stButton > button[kind="primary"]:hover,
.stButton > button[data-testid="stBaseButton-primary"]:hover{opacity:.88;}
.stButton > button[kind="secondary"],
.stButton > button[data-testid="stBaseButton-secondary"],
.stDownloadButton > button{
  background:var(--surface); color:var(--text); border:1px solid var(--border-strong);
}
.stButton > button[kind="secondary"]:hover,
.stDownloadButton > button:hover{background:var(--surface-2); border-color:var(--text-3);}
.stButton > button:focus-visible, .stDownloadButton > button:focus-visible{
  outline:2px solid var(--ink); outline-offset:2px;
}

/* ── File uploader ────────────────────────────────────────────────── */
[data-testid="stFileUploader"] section,
[data-testid="stFileUploaderDropzone"]{
  background:var(--surface-2); border:1px dashed var(--border-strong);
  border-radius:var(--r-panel); padding:26px 20px;
  transition:border-color var(--dur) var(--ease), background var(--dur) var(--ease);
}
[data-testid="stFileUploader"] section:hover,
[data-testid="stFileUploaderDropzone"]:hover{border-color:var(--text-3); background:var(--surface);}
[data-testid="stFileUploader"] small{color:var(--text-3);}
[data-testid="stFileUploader"] button{
  background:var(--surface)!important; color:var(--text)!important;
  border:1px solid var(--border-strong)!important; border-radius:var(--r-ctl)!important;
  font-weight:550!important; box-shadow:none!important;
}
[data-testid="stFileUploaderFile"]{font-size:13px;}

/* ── Inputs ───────────────────────────────────────────────────────── */
[data-baseweb="select"] > div, .stTextInput input, .stNumberInput input{
  border-radius:var(--r-ctl)!important; border-color:var(--border-strong)!important;
  background:var(--surface)!important; font-size:13.5px!important;
}
/* BaseWeb accents → ink, theme-aware (config.toml sets a static fallback but
   cannot follow light/dark; these do). Covers slider thumb + filled track,
   checkbox tick, radio dot and the spinner arc, which otherwise show the
   Streamlit stock red. */
.stSlider [data-baseweb="slider"] div[role="slider"]{
  border-color:var(--ink)!important; background:var(--ink)!important;
}
.stSlider [data-baseweb="slider"] [data-testid="stSliderTrack"] > div{background:var(--ink)!important;}
[data-baseweb="checkbox"] [data-testid="stCheckbox"] input:checked + div,
[data-testid="stCheckbox"] [data-baseweb="checkbox"] span[aria-checked="true"]{
  background:var(--ink)!important; border-color:var(--ink)!important;
}
[data-baseweb="radio"] div[aria-checked="true"]{border-color:var(--ink)!important;}
[data-baseweb="radio"] div[aria-checked="true"] > div{background:var(--ink)!important;}
[data-testid="stSpinner"] svg{color:var(--ink)!important; stroke:var(--ink)!important;}
label, .stRadio label, .stCheckbox label{font-size:13.5px!important; color:var(--text)!important;}
[data-testid="stWidgetLabel"] p{
  font-size:11px!important; font-weight:550!important; letter-spacing:.04em;
  text-transform:uppercase; color:var(--text-3)!important;
}

/* ── Expander ─────────────────────────────────────────────────────── */
[data-testid="stExpander"]{
  border:1px solid var(--border)!important; border-radius:var(--r-panel)!important;
  background:var(--surface)!important; box-shadow:none!important;
}
[data-testid="stExpander"] summary{font-size:13.5px!important; font-weight:550!important;}
[data-testid="stExpander"] summary:hover{color:var(--text)!important;}

/* ── Progress / spinner ───────────────────────────────────────────── */
[data-testid="stSpinner"]{color:var(--text-2)!important; font-size:13.5px;}
.stProgress > div > div > div{background:var(--ink)!important;}

/* Keyboard focus is always ink, never Streamlit's stock red ring — including
   on BaseWeb elements that carry their own [data-focus-visible] rule. */
:focus-visible,
[data-focus-visible]:focus-visible,
[data-baseweb] :focus-visible{
  outline:2px solid var(--ink)!important; outline-offset:2px!important;
  box-shadow:none!important;
}

@media (prefers-reduced-motion: reduce){
  *,*::before,*::after{
    animation-duration:1ms!important; transition-duration:1ms!important;
    animation-iteration-count:1!important;
  }
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

STATED_RE = re.compile(r"^\s*(\d[\d,]*)\s+Items?\s*-\s*\$([\d,]+\.\d{2})", re.I)


def esc(v) -> str:
    return (
        str("" if v is None else v)
        .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    )


def session_tmp() -> str:
    """One temp directory per browser session, emptied at the start of each run.

    Streamlit reruns the whole script on every interaction and `st.stop()`
    raises, so cleanup placed at the end of the script never runs. Clearing on
    entry is the only placement that cannot be skipped.
    """
    d = st.session_state.get("_tmp")
    if not d or not os.path.isdir(d):
        d = tempfile.mkdtemp(prefix="pdf2xl_")
        st.session_state["_tmp"] = d
    return d


def reset_tmp() -> str:
    d = session_tmp()
    for name in os.listdir(d):
        try:
            p = os.path.join(d, name)
            shutil.rmtree(p) if os.path.isdir(p) else os.unlink(p)
        except OSError:
            pass
    return d


def stated_totals(pdf_path: str, password: str = ""):
    """Read the report's own 'N Items - $X' header line, when present.

    This is what the team reconciles against by hand, so the interface should
    do it for them rather than making them open the PDF to check.
    """
    try:
        import pdfplumber
        with pdfplumber.open(pdf_path, password=password) as pdf:
            for page in pdf.pages[:2]:
                for line in (page.extract_text() or "").split("\n"):
                    m = STATED_RE.match(line.strip())
                    if m:
                        return (int(m.group(1).replace(",", "")),
                                float(m.group(2).replace(",", "")))
    except Exception:
        pass
    return None


def money(v) -> str:
    try:
        return f"${float(v):,.2f}"
    except (TypeError, ValueError):
        return "—"


def read_sheet(path: str, sheet: str, limit: int):
    """Return (header, rows, total_rows, amount_sum) for a written sheet."""
    wb = load_workbook(path, read_only=True)
    ws = wb[sheet]
    header, rows, total, amt = None, [], 0, 0.0
    for r in ws.iter_rows(values_only=True):
        cells = list(r)
        if not any(c not in (None, "") for c in cells):
            continue
        if header is None:
            # Skip a title row written above the header
            if sum(1 for c in cells if c not in (None, "")) == 1:
                continue
            header = [str(c) if c is not None else "" for c in cells]
            continue
        total += 1
        if "Amount" in header:
            v = cells[header.index("Amount")]
            if isinstance(v, (int, float)):
                amt += float(v)
        if len(rows) < limit:
            rows.append(cells)
    wb.close()
    return header or [], rows, total, amt


def ledger_html(header, rows, total_rows) -> str:
    num_cols = {"s. no.", "amount", "units", "qty", "quantity"}
    mono_cols = {"order no", "certificate no", "amount"}
    ths, cls = [], []
    for h in header:
        k = str(h).strip().lower()
        c = "n" if k in num_cols else ""
        cls.append((c, "m" if k in mono_cols else ""))
        ths.append(f'<th class="{c}">{esc(h)}</th>')
    trs = []
    for row in rows:
        tds = []
        for i, cell in enumerate(row):
            c, m = cls[i] if i < len(cls) else ("", "")
            val = money(cell) if (i < len(header)
                                  and str(header[i]).strip().lower() == "amount"
                                  and isinstance(cell, (int, float))) else esc(cell)
            tds.append(f'<td class="{c} {m}">{val}</td>')
        trs.append("<tr>" + "".join(tds) + "</tr>")
    shown = len(rows)
    foot = (f'<div class="ledger-foot num">Showing {shown} of {total_rows} rows</div>'
            if shown < total_rows else
            f'<div class="ledger-foot num">{total_rows} rows</div>')
    return (
        '<div class="ledger-wrap"><div class="ledger-scroll">'
        f'<table class="ledger"><thead><tr>{"".join(ths)}</tr></thead>'
        f'<tbody>{"".join(trs)}</tbody></table></div>{foot}</div>'
    )


cfg = load_config()


@st.cache_resource
def get_logger(log_dir: str):
    """One log file per server run, not one per script rerun.

    Streamlit re-executes this module top to bottom on every widget
    interaction, and setup_logging() mints a fresh timestamped file on each
    call. That littered logs/ with near-empty files and, worse, split a single
    conversion's records across several of them - so "check the newest log"
    could point at the wrong file. Caching the resource pins one logger for the
    life of the server.
    """
    return setup_logging(log_dir)


logger = get_logger(cfg.get("log_dir", "logs"))

# ─────────────────────────────────────────────────────────────────────────────
# MASTHEAD
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="app-mast">
  <div class="app-brand"><span class="mark">◆</span><span class="name">PDF to Excel</span></div>
  <div class="app-env">Converter</div>
</div>
<h1 class="app-h1">PDF to Excel</h1>
<p class="app-sub">Extracts every line item from a supplier PDF and reconciles the
result against the invoice's own stated total, so you can see the conversion is
complete before you download it. Drop a Word, Excel or PowerPoint document
instead and it is converted <em>into</em> PDF.</p>
""", unsafe_allow_html=True)

# The label is hidden but is still the accessible name a screen reader reads
# out, so it has to describe what the control now takes - not just PDFs.
uploaded = st.file_uploader("Report PDF, or a document to convert into PDF",
                            type=UPLOAD_TYPES, label_visibility="collapsed")

if not uploaded:
    st.markdown("""
    <p class="app-help"><strong>The file decides what happens</strong> — a PDF is read,
    anything else is converted into one. There is nothing to switch.</p>
    <p class="app-help">Digital, scanned and mixed PDFs are all handled, and the type is
    detected per page, so there is nothing to configure for a normal report. Scanned
    pages need Tesseract installed; without it they are skipped and reported rather
    than dropped.</p>
    <p class="app-help">Password-protected PDFs are read too. You are asked for the
    password when one is dropped; it is used for that conversion only and is never
    written to the log or to disk.</p>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENT
# ─────────────────────────────────────────────────────────────────────────────
tmp_dir = reset_tmp()
# Written with its real suffix: LibreOffice picks its import filter from the
# extension, so a .docx saved as "source.pdf" would be refused.
suffix = Path(uploaded.name).suffix.lower()
src_path = os.path.join(tmp_dir, "source" + suffix)
with open(src_path, "wb") as f:
    f.write(uploaded.getvalue())
size_kb = len(uploaded.getvalue()) / 1024
stem = os.path.splitext(uploaded.name)[0]

# ─────────────────────────────────────────────────────────────────────────────
# DOCUMENT → PDF   (a non-PDF was dropped)
# ─────────────────────────────────────────────────────────────────────────────
if suffix != ".pdf":
    engine_ready = libreoffice_present(cfg)

    st.markdown(f"""
    <h2 class="app-h2">Document</h2>
    <div class="app-metrics">
      <div><p class="app-label">Format</p>
        <div class="app-metric-v">{esc(suffix.lstrip('.').upper())}</div></div>
      <div><p class="app-label">Size</p>
        <div class="app-metric-v">{size_kb:,.0f} KB</div></div>
      <div><p class="app-label">Converts to</p>
        <div class="app-metric-v">PDF</div></div>
      <div><p class="app-label">Engine</p>
        <div class="app-metric-v">{'Ready' if engine_ready else 'Missing'}</div></div>
    </div>
    <p class="app-help mono">{esc(uploaded.name)}</p>
    """, unsafe_allow_html=True)

    # Explained in place, before the operator commits to an action that cannot
    # work. Deliberately not amber: amber means "rows excluded from the output"
    # and nothing else on either surface (DESIGN.md rule 1).
    if not engine_ready:
        st.markdown("""
        <div class="app-panel">
          <p class="app-label">LibreOffice required</p>
          <p class="app-help">Converting a document into PDF is done by a headless
          LibreOffice. It is separate software this tool does not bundle, and it is
          not installed here.</p>
          <p class="app-help">Install it, then reload this page — there is nothing
          else to configure. If it is installed somewhere unusual, set
          <span class="mono">libreoffice_path</span> in
          <span class="mono">config.json</span>.</p>
          <p class="app-help">Reading PDFs is unaffected and needs none of this.</p>
        </div>
        """, unsafe_allow_html=True)
        st.stop()

    if not st.button("Convert to PDF", type="primary"):
        st.stop()

    doc_out = os.path.join(tmp_dir, "out")
    os.makedirs(doc_out, exist_ok=True)
    with st.spinner("Converting with LibreOffice…"):
        conv = DocumentToPDFConverter(cfg, logger).convert(
            src_path, doc_out, output_name=stem)

    if not conv["success"]:
        st.markdown(f"""
        <h2 class="app-h2">Result</h2>
        <div class="notice fail"><span class="t">Conversion failed</span>
        {esc(conv['error'])}
        <span class="b">The full error is in the log under
        <span class="mono">{esc(cfg.get('log_dir', 'logs'))}/</span>.</span></div>
        """, unsafe_allow_html=True)
        st.stop()

    try:
        import pdfplumber
        with pdfplumber.open(conv["output"]) as _pdf:
            page_count = len(_pdf.pages)
    except Exception:
        page_count = 0

    out_kb = os.path.getsize(conv["output"]) / 1024
    st.markdown(f"""
    <h2 class="app-h2">Result</h2>
    <div class="app-panel">
      <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:14px;">
        <span class="pill ok"><span class="dot"></span>Converted</span>
        <span class="app-help mono" style="margin:0">{conv['duration_sec']:.1f}s</span>
      </div>
      <div class="verdict ok">
        <span class="verdict-fig">{page_count or '—'}</span>
        <span class="verdict-cap">page(s) written · {out_kb:,.0f} KB</span>
      </div>
      <p class="app-help" style="margin:0">The PDF was checked onto disk before this
      was shown — LibreOffice can exit successfully having written nothing, so the
      file itself is what confirms the conversion.</p>
    </div>
    """, unsafe_allow_html=True)

    with open(conv["output"], "rb") as fh:
        st.download_button(f"Download {stem}.pdf", fh.read(),
                           f"{stem}.pdf", "application/pdf")
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# ENCRYPTED PDF — ask, do not fail
# ─────────────────────────────────────────────────────────────────────────────
access = check_pdf_access(src_path)
if access == ACCESS_PASSWORD_REQUIRED:
    st.markdown(f"""
    <h2 class="app-h2">Document</h2>
    <div class="app-panel">
      <p class="app-label">Password required</p>
      <p class="app-help"><span class="mono">{esc(uploaded.name)}</span> is
      encrypted. Enter its password to read it — it is used for this conversion
      only, is never written to the log or to <span class="mono">config.json</span>,
      and is gone when you close the tab.</p>
    </div>
    """, unsafe_allow_html=True)

    pdf_password = st.text_input("PDF password", type="password",
                                 label_visibility="collapsed",
                                 placeholder="Password")
    if not pdf_password:
        st.stop()

    if check_pdf_access(src_path, pdf_password) != ACCESS_OK:
        st.markdown("""
        <div class="notice fail"><span class="t">Password incorrect</span>
        That password did not unlock the file.
        <span class="b">Check it and try again.</span></div>
        """, unsafe_allow_html=True)
        st.stop()

    cfg["pdf_password"] = pdf_password

try:
    info = PDFClassifier(logger, cfg.get("pdf_password")).classify(src_path)
except Exception:
    info = {"page_count": 0, "type": "unknown", "text_pages": [], "scanned_pages": []}

stated = stated_totals(src_path, cfg.get("pdf_password") or "")

st.markdown(f"""
<h2 class="app-h2">Document</h2>
<div class="app-metrics">
  <div><p class="app-label">Pages</p><div class="app-metric-v">{info['page_count']}</div></div>
  <div><p class="app-label">Type</p><div class="app-metric-v">{esc(info['type'].title())}</div></div>
  <div><p class="app-label">Text pages</p><div class="app-metric-v">{len(info['text_pages'])}</div></div>
  <div><p class="app-label">Scanned</p><div class="app-metric-v">{len(info['scanned_pages'])}</div></div>
</div>
<p class="app-help mono">{esc(uploaded.name)} · {size_kb:,.0f} KB{
  ' · states ' + str(stated[0]) + ' items, ' + money(stated[1]) if stated else ''
}</p>
""", unsafe_allow_html=True)

with st.expander("Conversion settings"):
    c1, c2 = st.columns(2)
    with c1:
        output_fmt = st.radio("Output format", ["Excel (.xlsx)", "CSV (per table)"],
                              help="Excel gives one workbook with a sheet per dataset. "
                                   "CSV gives one file per extracted table.")
        method = st.selectbox(
            "Extraction engine", ["auto", "pdfplumber", "camelot", "tabula"],
            format_func=lambda x: {
                "auto": "Auto — recommended",
                "pdfplumber": "pdfplumber — fast, digital PDFs",
                "camelot": "Camelot — ruled tables (needs Ghostscript)",
                "tabula": "Tabula — fallback (needs Java)",
            }[x],
            help="Auto tries each engine in order and stops at the first that finds "
                 "tables. Change this only if Auto misses a table.")
        remove_empty = st.checkbox("Remove empty rows and columns", value=True)
    with c2:
        ocr_lang = st.selectbox(
            "OCR language", ["eng", "fra", "deu", "spa", "chi_sim", "jpn"],
            format_func=lambda x: {"eng": "English", "fra": "French", "deu": "German",
                                   "spa": "Spanish", "chi_sim": "Chinese (Simplified)",
                                   "jpn": "Japanese"}[x],
            help="Applies only to scanned pages.")
        ocr_dpi = st.slider("OCR resolution (DPI)", 72, 600, 300, 50,
                            help="300 suits most scans. Higher is slower but reads "
                                 "small print better.")
        add_meta = st.checkbox("Include metadata sheet", value=True,
                               help="Adds a _Metadata sheet recording the source file, "
                                    "counts and any excluded rows.")
    preview_rows = st.slider("Preview rows", 0, 50, 12)

run = st.button("Convert", type="primary")
if not run:
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# CONVERT
# ─────────────────────────────────────────────────────────────────────────────
cfg.update({
    "table_extraction_method": method,
    "ocr_language": ocr_lang,
    "ocr_dpi": ocr_dpi,
    "remove_empty_rows": remove_empty,
    "remove_empty_cols": remove_empty,
    "add_metadata_sheet": add_meta,
})

out_dir = os.path.join(tmp_dir, "out")
os.makedirs(out_dir, exist_ok=True)
start = time.time()

with st.spinner("Reading the document…"):
    if "CSV" in output_fmt:
        export = export_to_csv(src_path, out_dir, cfg, logger)
        files = export["files"]
        csv_unparsed = export["unparsed_rows"]
        duration = time.time() - start
        if not files:
            st.markdown("""
            <div class="notice fail"><span class="t">No data found</span>
            Nothing in this PDF matched a table or a readable report layout.
            <span class="b">Try a different engine under Conversion settings, or check
            the pages actually contain tabular data.</span></div>
            """, unsafe_allow_html=True)
            st.stop()
        csv_pill = ('<span class="pill held"><span class="dot"></span>Rows excluded</span>'
                    if csv_unparsed else
                    '<span class="pill ok"><span class="dot"></span>Complete</span>')
        st.markdown(f"""
        <h2 class="app-h2">Result</h2>
        <div class="app-panel">{csv_pill}
        <p class="app-help">{len(files)} CSV file(s) in {duration:.1f}s.</p></div>
        """, unsafe_allow_html=True)
        # The excluded-row warning belongs on every output path, not just the
        # Excel one - a CSV export that quietly omits rows is the exact failure
        # this tool exists to prevent.
        if csv_unparsed:
            st.markdown(f"""
            <div class="notice held"><span class="t">{csv_unparsed} row(s) excluded from this export</span>
            They did not match the expected format and are <strong>not</strong> in the CSV.
            <span class="b">Check them before using this data — each excluded row is written in
            full to the newest log in
            <span class="mono">{esc(cfg.get('log_dir', 'logs'))}/</span>.</span></div>
            """, unsafe_allow_html=True)
        for path in files:
            with open(path, "rb") as fh:
                st.download_button(f"Download {os.path.basename(path)}", fh.read(),
                                   os.path.basename(path), "text/csv", key=path)
        for path in files[:2]:
            if preview_rows:
                st.markdown(f'<p class="app-label" style="margin-top:18px">'
                            f'{esc(os.path.basename(path))}</p>', unsafe_allow_html=True)
                df = pd.read_csv(path).head(preview_rows)
                st.markdown(ledger_html(list(df.columns), df.values.tolist(), len(df)),
                            unsafe_allow_html=True)
        st.stop()

    xlsx = os.path.join(out_dir, f"{stem}_{datetime.now():%Y%m%d}.xlsx")
    result = PDFProcessor(cfg, logger).process(src_path, xlsx)
    duration = time.time() - start

if not result["success"]:
    st.markdown(f"""
    <h2 class="app-h2">Result</h2>
    <div class="notice fail"><span class="t">Conversion failed</span>
    {esc(result['error'])}
    <span class="b">The full error is in the log under
    <span class="mono">{esc(cfg.get('log_dir', 'logs'))}/</span>.</span></div>
    """, unsafe_allow_html=True)
    st.stop()

# ─────────────────────────────────────────────────────────────────────────────
# RESULT — reconciliation first, download last
# ─────────────────────────────────────────────────────────────────────────────
unparsed = result.get("unparsed_rows", 0)
rows_out = result.get("rows_extracted", 0)

_wb = load_workbook(xlsx, read_only=True)
try:
    sheets = [s for s in _wb.sheetnames if not s.startswith("_")]
finally:
    # read_only leaves a file handle open, which locks the file on Windows.
    _wb.close()

# Reconcile against *every* data sheet, not just the first. A workbook split
# across several sheets otherwise compared the invoice's stated total against a
# fraction of the extraction and reported a shortfall that did not exist.
header, rows, first_total = [], [], 0
total_rows, amount_sum = 0, 0.0
for _idx, _name in enumerate(sheets):
    try:
        _h, _r, _t, _a = read_sheet(
            xlsx, _name, (preview_rows or 12) if _idx == 0 else 0)
    except Exception:
        continue
    if _idx == 0:
        header, rows, first_total = _h, _r, _t
    total_rows += _t
    amount_sum += _a

reconciles = bool(stated) and stated[0] == total_rows and abs(amount_sum - stated[1]) < 0.005
if unparsed:
    pill = '<span class="pill held"><span class="dot"></span>Rows excluded</span>'
elif stated and reconciles:
    pill = '<span class="pill ok"><span class="dot"></span>Reconciled</span>'
else:
    pill = '<span class="pill ok"><span class="dot"></span>Complete</span>'

# Lead with the verdict at metric scale — this figure is the reason the screen
# exists — then show the arithmetic that produced it underneath.
if stated:
    d_rows, d_amt = total_rows - stated[0], amount_sum - stated[1]
    if reconciles and not unparsed:
        verdict = ('<div class="verdict ok"><span class="verdict-fig">$0.00</span>'
                   '<span class="verdict-cap">reconciles exactly to the invoice total</span></div>')
    else:
        amt = "$0.00" if abs(d_amt) < 0.005 else f"{'-' if d_amt < 0 else '+'}${abs(d_amt):,.2f}"
        verdict = (f'<div class="verdict held"><span class="verdict-fig">{amt}</span>'
                   f'<span class="verdict-cap">{d_rows:+d} rows against the invoice total — '
                   f'reconcile before use</span></div>')
    recon_body = f"""{verdict}
      <div class="recon">
        <div class="k">Stated on the invoice</div>
        <div class="v">{stated[0]}</div><div class="v">{money(stated[1])}</div>
        <div class="k">Extracted</div>
        <div class="v">{total_rows}</div><div class="v">{money(amount_sum)}</div>
      </div>"""
else:
    verdict = (f'<div class="verdict ok"><span class="verdict-fig">{rows_out}</span>'
               f'<span class="verdict-cap">rows extracted across '
               f'{result["tables_found"]} dataset(s)</span></div>')
    recon_body = f"""{verdict}
      <p class="app-help" style="margin:0">This report states no total to reconcile against;
      compare the sheet below against the source before relying on it.</p>"""

# Heading and panel share one block, so the heading-to-panel spacing is the
# heading's own margin-bottom, not that margin plus the inter-block gap.
st.markdown(f"""
<h2 class="app-h2">Result</h2>
<div class="app-panel">
  <div style="display:flex;align-items:center;justify-content:space-between;gap:12px;margin-bottom:14px;">
    {pill}
    <span class="app-help mono" style="margin:0">{result['pages']} pages · {duration:.1f}s</span>
  </div>
  {recon_body}
</div>
""", unsafe_allow_html=True)

if unparsed:
    st.markdown(f"""
    <div class="notice held"><span class="t">{unparsed} row(s) excluded from this file</span>
    They did not match the expected format and are <strong>not</strong> in the workbook.
    <span class="b">Check them before using this data — each excluded row is written in
    full to the newest log in
    <span class="mono">{esc(cfg.get('log_dir', 'logs'))}/</span>.</span></div>
    """, unsafe_allow_html=True)
elif stated and not reconciles:
    st.markdown("""
    <div class="notice held"><span class="t">Does not reconcile to the invoice total</span>
    No rows were rejected, but the extracted figures differ from what the invoice states.
    <span class="b">Compare the sheet against the PDF before relying on it.</span></div>
    """, unsafe_allow_html=True)

with open(xlsx, "rb") as fh:
    st.download_button(
        f"Download {stem}.xlsx", fh.read(), f"{stem}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if preview_rows and header:
    # Preview state is kept separate from the reconciliation totals above, which
    # span every sheet. Reusing those variables here made the chosen sheet's row
    # count overwrite the figure the verdict was computed from.
    sel = sheets[0]
    sel_header, sel_rows, sel_total = header, rows, first_total
    if len(sheets) > 1:
        sel = st.selectbox("Sheet", sheets)
        try:
            sel_header, sel_rows, sel_total, _ = read_sheet(xlsx, sel, preview_rows)
        except Exception:
            pass
    st.markdown(f'<h2 class="app-h2">{esc(sel)}</h2>', unsafe_allow_html=True)
    st.markdown(ledger_html(sel_header, sel_rows, sel_total), unsafe_allow_html=True)
